import hashlib
import hmac
import io
import json
import logging
import os
import re
import subprocess
import sys
import threading
import base64
import tempfile
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import psycopg
import requests
from flask import Flask, jsonify, request
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageOps


logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))
log = logging.getLogger("porsline-reporter")

app = Flask(__name__)
RUN_LOCK = threading.Lock()
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "assets" / "report-template.xlsx.b64"
CERTIFICATE_TEMPLATE_PATH = BASE_DIR / "assets" / "certificate-template.pptx"
CERTIFICATE_WORKER_PATH = BASE_DIR / "certificate_worker.py"

PORSLINE_BASE_URL = os.getenv("PORSLINE_BASE_URL", "https://survey.porsline.ir").rstrip("/")
PORSLINE_API_KEY = os.getenv("PORSLINE_API_KEY", "")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")
DATABASE_URL = os.getenv("DATABASE_URL", "")
INJECTION_SURVEY_CODE = os.getenv("INJECTION_SURVEY_CODE", "6Hf5AK7g")
TECHNICIAN_SURVEY_CODE = os.getenv("TECHNICIAN_SURVEY_CODE", "jiUT4eKo")
TAHER_INJECTION_SURVEY_CODE = os.getenv("TAHER_INJECTION_SURVEY_CODE", "mobh0bQS")
TAHER_SUTURE_SURVEY_CODE = os.getenv("TAHER_SUTURE_SURVEY_CODE", "sNUa7F2D")
ZOHIRI_SURVEY_CODE = os.getenv("ZOHIRI_SURVEY_CODE", "ox2HIlC4")
REPORT_PREFIX = os.getenv("REPORT_PREFIX", "تیر-تکنسین")
APP_SECRET = os.getenv("APP_SECRET", "")
TELEGRAM_POLLING_ENABLED = os.getenv("TELEGRAM_POLLING_ENABLED", "true").lower() == "true"
WEBHOOK_BASE_URL = os.getenv("WEBHOOK_BASE_URL", "").rstrip("/")
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "")
BOT_ACCESS_CODE = os.getenv("BOT_ACCESS_CODE", "")
DISPLAY_TIMEZONE = os.getenv("DISPLAY_TIMEZONE", "Asia/Tehran")

TEACHER_COLORS = {
    "white": ("سفید", "FFFFFF"),
    "yellow": ("زرد", "FFF2CC"),
    "green": ("سبز", "C6EFCE"),
    "red": ("قرمز", "FFC7CE"),
    "blue": ("آبی", "BDD7EE"),
    "orange": ("نارنجی", "F4B183"),
    "purple": ("بنفش", "D9D2E9"),
    "gray": ("خاکستری", "D9D9D9"),
    "pink": ("صورتی", "F4CCCC"),
    "cyan": ("فیروزه‌ای", "DDEBF7"),
    "lime": ("لیمویی", "E2F0D9"),
    "gold": ("طلایی", "FFD966"),
    "lavender": ("یاسی", "E4DFEC"),
    "peach": ("هلویی", "FCE4D6"),
    "mint": ("نعنایی", "D0EAD0"),
    "sky": ("آسمانی", "C9DAF8"),
    "rose": ("رز", "EAD1DC"),
    "sand": ("کرم", "E6D5B8"),
    "teal": ("سبزآبی", "A2D9CE"),
    "plum": ("آلویی", "C9B2D9"),
}

FIVE_REPORTS = [
    (INJECTION_SURVEY_CODE, "تزریقات"),
    (TECHNICIAN_SURVEY_CODE, "تکنسین داروخانه"),
    (TAHER_INJECTION_SURVEY_CODE, "تزریقات خانم طاهرخانی"),
    (TAHER_SUTURE_SURVEY_CODE, "بخیه خانم طاهرخانی"),
    (ZOHIRI_SURVEY_CODE, "مدارک خانم ظهیری"),
]

SINGLE_REPORT_COMMANDS = {
    "/zanyar_t": (INJECTION_SURVEY_CODE, "تزریقات"),
    "/zanyar_tek": (TECHNICIAN_SURVEY_CODE, "تکنسین داروخانه"),
    "/taher_t": (TAHER_INJECTION_SURVEY_CODE, "تزریقات خانم طاهرخانی"),
    "/taher_b": (TAHER_SUTURE_SURVEY_CODE, "بخیه خانم طاهرخانی"),
}

UTILITY_COMMANDS = {
    "/help", "/status", "/report", "/report_all", "/confirm_report_all",
    "/cancel_report", "/report_all_new", "/add_form", "/remove_form", "/forms",
}


def require_settings():
    missing = []
    for key, value in {
        "PORSLINE_API_KEY": PORSLINE_API_KEY,
        "TELEGRAM_BOT_TOKEN": TELEGRAM_BOT_TOKEN,
        "DATABASE_URL": DATABASE_URL,
    }.items():
        if not value:
            missing.append(key)
    if missing:
        raise RuntimeError("Missing settings: " + ", ".join(missing))


def db():
    return psycopg.connect(DATABASE_URL)


def init_db():
    require_settings()
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS processed_responses (
                survey_code TEXT NOT NULL,
                response_key TEXT NOT NULL,
                processed_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                PRIMARY KEY (survey_code, response_key)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS app_state (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS response_send_history (
                id BIGSERIAL PRIMARY KEY,
                survey_code TEXT NOT NULL,
                response_key TEXT NOT NULL,
                sent_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cur.execute(
            "CREATE INDEX IF NOT EXISTS response_send_history_lookup "
            "ON response_send_history(survey_code, response_key, sent_at DESC)"
        )
        cur.execute(
            """
            INSERT INTO response_send_history(survey_code, response_key, sent_at)
            SELECT p.survey_code, p.response_key, p.processed_at
            FROM processed_responses p
            WHERE NOT EXISTS (
                SELECT 1 FROM response_send_history h
                WHERE h.survey_code=p.survey_code AND h.response_key=p.response_key
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS custom_forms (
                bot_command TEXT PRIMARY KEY,
                survey_code TEXT NOT NULL UNIQUE,
                report_name TEXT NOT NULL,
                created_by TEXT NOT NULL,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS teachers (
                name TEXT PRIMARY KEY,
                created_by TEXT NOT NULL,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cur.execute("ALTER TABLE teachers ADD COLUMN IF NOT EXISTS active BOOLEAN NOT NULL DEFAULT TRUE")
        cur.execute("ALTER TABLE teachers ADD COLUMN IF NOT EXISTS color_key TEXT")
        cur.execute("ALTER TABLE custom_forms ADD COLUMN IF NOT EXISTS teacher_name TEXT")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS builtin_forms (
                survey_code TEXT PRIMARY KEY,
                report_name TEXT NOT NULL,
                teacher_name TEXT NOT NULL,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        for table in ("custom_forms", "builtin_forms"):
            cur.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS certificate_type TEXT NOT NULL DEFAULT 'nursing'")
            cur.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS course_title TEXT")
            cur.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS course_duration TEXT")
            cur.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS certificate_instructor TEXT")
            cur.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS certificate_venue TEXT")
            cur.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS certificate_organization TEXT")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS report_batches (
                id BIGSERIAL PRIMARY KEY,
                short_id TEXT UNIQUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                report_names TEXT NOT NULL,
                empty_report_names TEXT NOT NULL DEFAULT '[]',
                row_count INTEGER NOT NULL,
                first_name TEXT NOT NULL DEFAULT '',
                last_name TEXT NOT NULL DEFAULT '',
                is_combined BOOLEAN NOT NULL DEFAULT FALSE,
                nursing_received_at TIMESTAMPTZ,
                posted_at TIMESTAMPTZ,
                day_40_warning_sent_at TIMESTAMPTZ
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS report_batch_items (
                id BIGSERIAL PRIMARY KEY,
                batch_id BIGINT NOT NULL REFERENCES report_batches(id) ON DELETE CASCADE,
                survey_code TEXT NOT NULL,
                response_key TEXT NOT NULL,
                report_name TEXT NOT NULL,
                teacher_name TEXT NOT NULL,
                color_key TEXT NOT NULL DEFAULT 'white',
                persian_name TEXT NOT NULL DEFAULT '',
                english_name TEXT NOT NULL DEFAULT '',
                national_id TEXT NOT NULL DEFAULT '',
                row_position INTEGER NOT NULL,
                UNIQUE(batch_id, survey_code, response_key)
            )
            """
        )
        cur.execute("ALTER TABLE report_batch_items ADD COLUMN IF NOT EXISTS color_key TEXT NOT NULL DEFAULT 'white'")
        cur.execute("CREATE INDEX IF NOT EXISTS report_batch_items_person ON report_batch_items(national_id, persian_name)")
        cur.execute("CREATE INDEX IF NOT EXISTS report_batches_status ON report_batches(created_at DESC, nursing_received_at, posted_at)")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS self_certificate_batches (
                id BIGSERIAL PRIMARY KEY,
                short_id TEXT UNIQUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                report_names TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                output_mode TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'sent_to_printer'
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS self_certificate_items (
                id BIGSERIAL PRIMARY KEY,
                batch_id BIGINT NOT NULL REFERENCES self_certificate_batches(id) ON DELETE CASCADE,
                survey_code TEXT NOT NULL,
                response_key TEXT NOT NULL,
                report_name TEXT NOT NULL,
                teacher_name TEXT NOT NULL,
                persian_name TEXT NOT NULL,
                national_id TEXT NOT NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                UNIQUE(survey_code, response_key)
            )
            """
        )
        cur.execute("CREATE INDEX IF NOT EXISTS self_certificate_person ON self_certificate_items(national_id, persian_name)")
        cur.execute("SELECT value FROM app_state WHERE key='teacher_catalog_seeded'")
        if not cur.fetchone():
            cur.executemany(
                """
                INSERT INTO teachers(name, created_by) VALUES(%s, %s)
                ON CONFLICT(name) DO NOTHING
                """,
                [("اسکالپل", "system"), ("خانم طاهرخانی", "system"), ("خانم ظهیری", "system")],
            )
            cur.executemany(
                """
                INSERT INTO builtin_forms(survey_code, report_name, teacher_name)
                VALUES(%s, %s, %s) ON CONFLICT(survey_code) DO NOTHING
                """,
                [
                    (INJECTION_SURVEY_CODE, "تزریقات", "اسکالپل"),
                    (TECHNICIAN_SURVEY_CODE, "تکنسین داروخانه", "اسکالپل"),
                    (TAHER_INJECTION_SURVEY_CODE, "تزریقات خانم طاهرخانی", "خانم طاهرخانی"),
                    (TAHER_SUTURE_SURVEY_CODE, "بخیه خانم طاهرخانی", "خانم طاهرخانی"),
                    (ZOHIRI_SURVEY_CODE, "مدارک خانم ظهیری", "خانم ظهیری"),
                ],
            )
            cur.execute(
                """
                INSERT INTO app_state(key, value) VALUES('teacher_catalog_seeded', 'true')
                ON CONFLICT(key) DO NOTHING
                """
            )
            cur.execute(
                "UPDATE custom_forms SET teacher_name=%s WHERE teacher_name IS NULL OR teacher_name=''",
                ("اسکالپل",),
            )
        cur.execute("UPDATE teachers SET color_key='white' WHERE name='اسکالپل' AND color_key IS NULL")
        cur.execute("UPDATE teachers SET color_key='red' WHERE name='خانم طاهرخانی' AND color_key IS NULL")
        cur.execute("UPDATE teachers SET color_key='green' WHERE name='خانم ظهیری' AND color_key IS NULL")
        cur.execute("UPDATE teachers SET color_key='yellow' WHERE name='آقای حمیدی' AND color_key IS NULL")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS teachers_unique_color ON teachers(color_key) WHERE color_key IS NOT NULL")
        conn.commit()


def get_state(key, default=None):
    with db() as conn, conn.cursor() as cur:
        cur.execute("SELECT value FROM app_state WHERE key=%s", (key,))
        row = cur.fetchone()
    return row[0] if row else default


def set_state(key, value):
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO app_state(key, value) VALUES(%s, %s)
            ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value
            """,
            (key, str(value)),
        )
        conn.commit()


def get_custom_forms(active_only=True, teacher_name=None):
    query = "SELECT bot_command, survey_code, report_name FROM custom_forms"
    conditions = []
    params = []
    if active_only:
        conditions.append("active=TRUE")
    if teacher_name is not None:
        conditions.append("teacher_name=%s")
        params.append(teacher_name)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY created_at"
    with db() as conn, conn.cursor() as cur:
        cur.execute(query, params)
        return cur.fetchall()


def get_teachers(active_only=True):
    with db() as conn, conn.cursor() as cur:
        where = " WHERE active=TRUE" if active_only else ""
        cur.execute(
            """
            SELECT name FROM teachers
            """ + where + """
            ORDER BY CASE name
                WHEN 'اسکالپل' THEN 1
                WHEN 'خانم طاهرخانی' THEN 2
                WHEN 'خانم ظهیری' THEN 3
                ELSE 4
            END, created_at, name
            """
        )
        return [row[0] for row in cur.fetchall()]


def get_inactive_teachers():
    with db() as conn, conn.cursor() as cur:
        cur.execute("SELECT name FROM teachers WHERE active=FALSE ORDER BY created_at, name")
        return [row[0] for row in cur.fetchall()]


def get_teacher_colors():
    with db() as conn, conn.cursor() as cur:
        cur.execute("SELECT name, color_key FROM teachers WHERE active=TRUE ORDER BY created_at, name")
        return {name: color_key for name, color_key in cur.fetchall()}


def available_teacher_colors(exclude_teacher=None):
    with db() as conn, conn.cursor() as cur:
        if exclude_teacher:
            cur.execute("SELECT color_key FROM teachers WHERE color_key IS NOT NULL AND name<>%s", (exclude_teacher,))
        else:
            cur.execute("SELECT color_key FROM teachers WHERE color_key IS NOT NULL")
        used = {row[0] for row in cur.fetchall()}
    return [(key, label, rgb) for key, (label, rgb) in TEACHER_COLORS.items() if key not in used]


def set_teacher_color(teacher_name, color_key):
    if color_key not in TEACHER_COLORS:
        raise ValueError("⚠️ رنگ انتخاب‌شده معتبر نیست.")
    with db() as conn, conn.cursor() as cur:
        cur.execute("SELECT name FROM teachers WHERE color_key=%s AND name<>%s", (color_key, teacher_name))
        owner = cur.fetchone()
        if owner:
            raise ValueError(f"⚠️ این رنگ قبلاً برای مدرس «{owner[0]}» ثبت شده است.")
        cur.execute("UPDATE teachers SET color_key=%s WHERE name=%s AND active=TRUE", (color_key, teacher_name))
        if cur.rowcount != 1:
            raise ValueError("⚠️ مدرس فعال موردنظر پیدا نشد.")
        conn.commit()
    return TEACHER_COLORS[color_key][0]


def save_teacher(name, user_id, color_key=None):
    teacher_name = str(name or "").strip()
    if not teacher_name or len(teacher_name) > 60:
        raise ValueError("⚠️ نام مدرس باید بین ۱ تا ۶۰ کاراکتر باشد.")
    if color_key is not None and color_key not in TEACHER_COLORS:
        raise ValueError("⚠️ رنگ مدرس معتبر نیست.")
    with db() as conn, conn.cursor() as cur:
        cur.execute("SELECT 1 FROM teachers WHERE name=%s", (teacher_name,))
        if cur.fetchone():
            raise ValueError("⚠️ این مدرس قبلاً ثبت شده است؛ از فهرست مدرس‌ها انتخابش کنید.")
        cur.execute(
            "INSERT INTO teachers(name, created_by, color_key) VALUES(%s, %s, %s)",
            (teacher_name, str(user_id), color_key),
        )
        conn.commit()
    return teacher_name


def get_form_records(active_only=True, teacher_name=None):
    conditions = []
    params = []
    if active_only:
        conditions.append("active=TRUE")
    if teacher_name is not None:
        conditions.append("teacher_name=%s")
        params.append(teacher_name)
    where = " WHERE " + " AND ".join(conditions) if conditions else ""
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT 'builtin', survey_code, report_name, teacher_name FROM builtin_forms"
            + where + " ORDER BY created_at",
            params,
        )
        builtin = cur.fetchall()
        cur.execute(
            "SELECT 'custom', survey_code, report_name, teacher_name FROM custom_forms"
            + where + " ORDER BY created_at",
            params,
        )
        custom = cur.fetchall()
    return builtin + custom


def get_form_settings(survey_code):
    columns = (
        "certificate_type, course_title, course_duration, certificate_instructor, "
        "certificate_venue, certificate_organization"
    )
    with db() as conn, conn.cursor() as cur:
        for source, table in (("builtin", "builtin_forms"), ("custom", "custom_forms")):
            cur.execute(
                f"SELECT report_name, teacher_name, {columns} FROM {table} WHERE survey_code=%s",
                (survey_code,),
            )
            row = cur.fetchone()
            if row:
                return {
                    "source": source, "survey_code": str(survey_code),
                    "report_name": row[0], "teacher_name": row[1],
                    "certificate_type": row[2] or "nursing", "course_title": row[3] or "",
                    "course_duration": row[4] or "", "certificate_instructor": row[5] or "",
                    "certificate_venue": row[6] or "", "certificate_organization": row[7] or "",
                }
    return None


def update_form_certificate_settings(source, survey_code, settings):
    table = "builtin_forms" if source == "builtin" else "custom_forms"
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            f"""UPDATE {table} SET certificate_type=%s, course_title=%s,
                course_duration=%s, certificate_instructor=%s, certificate_venue=%s,
                certificate_organization=%s WHERE survey_code=%s""",
            (
                settings.get("certificate_type", "nursing"), settings.get("course_title"),
                settings.get("course_duration"), settings.get("certificate_instructor"),
                settings.get("certificate_venue"), settings.get("certificate_organization"), survey_code,
            ),
        )
        if cur.rowcount != 1:
            raise ValueError("⚠️ فرم موردنظر پیدا نشد.")
        conn.commit()


def get_inactive_form_records():
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT 'builtin', survey_code, report_name, teacher_name FROM builtin_forms "
            "WHERE active=FALSE ORDER BY created_at"
        )
        builtin = cur.fetchall()
        cur.execute(
            "SELECT 'custom', survey_code, report_name, teacher_name FROM custom_forms "
            "WHERE active=FALSE ORDER BY created_at"
        )
        custom = cur.fetchall()
    return builtin + custom


def rename_form(source, survey_code, new_name):
    report_name = str(new_name or "").strip()
    if not report_name or len(report_name) > 80:
        raise ValueError("⚠️ نام فرم باید بین ۱ تا ۸۰ کاراکتر باشد.")
    table = "builtin_forms" if source == "builtin" else "custom_forms"
    with db() as conn, conn.cursor() as cur:
        cur.execute(f"UPDATE {table} SET report_name=%s WHERE survey_code=%s AND active=TRUE", (report_name, survey_code))
        if cur.rowcount != 1:
            raise ValueError("⚠️ فرم فعال موردنظر پیدا نشد.")
        conn.commit()
    return report_name


def deactivate_form(source, survey_code):
    table = "builtin_forms" if source == "builtin" else "custom_forms"
    with db() as conn, conn.cursor() as cur:
        cur.execute(f"UPDATE {table} SET active=FALSE WHERE survey_code=%s AND active=TRUE", (survey_code,))
        changed = cur.rowcount == 1
        conn.commit()
    return changed


def rename_teacher(old_name, new_name):
    updated_name = str(new_name or "").strip()
    if not updated_name or len(updated_name) > 60:
        raise ValueError("⚠️ نام مدرس باید بین ۱ تا ۶۰ کاراکتر باشد.")
    with db() as conn, conn.cursor() as cur:
        cur.execute("SELECT 1 FROM teachers WHERE name=%s", (updated_name,))
        if cur.fetchone():
            raise ValueError("⚠️ مدرس دیگری با این نام وجود دارد.")
        cur.execute("UPDATE builtin_forms SET teacher_name=%s WHERE teacher_name=%s", (updated_name, old_name))
        cur.execute("UPDATE custom_forms SET teacher_name=%s WHERE teacher_name=%s", (updated_name, old_name))
        cur.execute("UPDATE teachers SET name=%s WHERE name=%s AND active=TRUE", (updated_name, old_name))
        if cur.rowcount != 1:
            raise ValueError("⚠️ مدرس فعال موردنظر پیدا نشد.")
        conn.commit()
    return updated_name


def deactivate_teacher(teacher_name):
    with db() as conn, conn.cursor() as cur:
        cur.execute("UPDATE teachers SET active=FALSE WHERE name=%s AND active=TRUE", (teacher_name,))
        if cur.rowcount != 1:
            return False
        cur.execute("UPDATE builtin_forms SET active=FALSE WHERE teacher_name=%s", (teacher_name,))
        cur.execute("UPDATE custom_forms SET active=FALSE WHERE teacher_name=%s", (teacher_name,))
        conn.commit()
    return True


def restore_form(source, survey_code, teacher_name):
    table = "builtin_forms" if source == "builtin" else "custom_forms"
    with db() as conn, conn.cursor() as cur:
        cur.execute("UPDATE teachers SET active=TRUE WHERE name=%s", (teacher_name,))
        cur.execute(f"UPDATE {table} SET active=TRUE WHERE survey_code=%s AND active=FALSE", (survey_code,))
        changed = cur.rowcount == 1
        conn.commit()
    return changed


def restore_teacher(teacher_name):
    with db() as conn, conn.cursor() as cur:
        cur.execute("UPDATE teachers SET active=TRUE WHERE name=%s AND active=FALSE", (teacher_name,))
        if cur.rowcount != 1:
            return False
        cur.execute("UPDATE builtin_forms SET active=TRUE WHERE teacher_name=%s", (teacher_name,))
        cur.execute("UPDATE custom_forms SET active=TRUE WHERE teacher_name=%s", (teacher_name,))
        conn.commit()
    return True


def move_form(source, survey_code, teacher_name):
    if teacher_name not in get_teachers():
        raise ValueError("⚠️ مدرس مقصد فعال نیست.")
    table = "builtin_forms" if source == "builtin" else "custom_forms"
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            f"UPDATE {table} SET teacher_name=%s WHERE survey_code=%s AND active=TRUE",
            (teacher_name, survey_code),
        )
        changed = cur.rowcount == 1
        conn.commit()
    return changed


def get_report_for_command(command):
    if command in SINGLE_REPORT_COMMANDS:
        survey_code = SINGLE_REPORT_COMMANDS[command][0]
        with db() as conn, conn.cursor() as cur:
            cur.execute(
                "SELECT survey_code, report_name FROM builtin_forms WHERE survey_code=%s AND active=TRUE",
                (survey_code,),
            )
            row = cur.fetchone()
        return tuple(row) if row else None
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT survey_code, report_name FROM custom_forms WHERE bot_command=%s AND active=TRUE",
            (command,),
        )
        row = cur.fetchone()
    return tuple(row) if row else None


def all_report_definitions(certificate_type=None):
    definitions = []
    for _source, survey_code, report_name, _teacher in get_form_records():
        if certificate_type:
            settings = get_form_settings(survey_code)
            if settings and settings["certificate_type"] != certificate_type:
                continue
        definitions.append((survey_code, report_name))
    return definitions


def custom_form_command(survey_code):
    digest = hashlib.sha256(str(survey_code).encode("utf-8")).hexdigest()[:16]
    return f"/form_{digest}"


def custom_form_button_label(report_name, survey_code):
    # Telegram limits reply-keyboard labels; the code suffix also makes names unique.
    suffix = str(survey_code)[-8:]
    max_name_length = max(1, 58 - len(suffix))
    return f"📋 {str(report_name)[:max_name_length]} • {suffix}"


def save_custom_form(command, survey_code, report_name, user_id, teacher_name="اسکالپل", settings=None):
    settings = settings or {"certificate_type": "nursing"}
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT bot_command, survey_code, active FROM custom_forms WHERE survey_code=%s OR bot_command=%s",
            (survey_code, command),
        )
        existing = cur.fetchone()
        if existing:
            old_command, old_code, active = existing
            if not active and old_command == command and old_code == survey_code:
                cur.execute(
                    """
                    UPDATE custom_forms
                    SET active=TRUE, report_name=%s, created_by=%s, teacher_name=%s,
                        certificate_type=%s, course_title=%s, course_duration=%s,
                        certificate_instructor=%s, certificate_venue=%s, certificate_organization=%s
                    WHERE bot_command=%s
                    """,
                    (report_name, user_id, teacher_name, settings.get("certificate_type", "nursing"),
                     settings.get("course_title"), settings.get("course_duration"),
                     settings.get("certificate_instructor"), settings.get("certificate_venue"),
                     settings.get("certificate_organization"), command),
                )
                conn.commit()
                return
            raise ValueError("این لینک یا دستور قبلاً ثبت شده است.")
        cur.execute(
            """
            INSERT INTO custom_forms(bot_command, survey_code, report_name, created_by, teacher_name,
                certificate_type, course_title, course_duration, certificate_instructor,
                certificate_venue, certificate_organization)
            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (command, survey_code, report_name, user_id, teacher_name,
             settings.get("certificate_type", "nursing"), settings.get("course_title"),
             settings.get("course_duration"), settings.get("certificate_instructor"),
             settings.get("certificate_venue"), settings.get("certificate_organization")),
        )
        conn.commit()


def deactivate_custom_form(command):
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            "UPDATE custom_forms SET active=FALSE WHERE bot_command=%s AND active=TRUE",
            (command,),
        )
        changed = cur.rowcount > 0
        conn.commit()
    return changed


def porsline_get(path, params=None):
    response = requests.get(
        f"{PORSLINE_BASE_URL}{path}",
        params=params,
        headers={
            "Authorization": f"API-Key {PORSLINE_API_KEY}",
            "Content-Type": "application/json",
        },
        timeout=90,
    )
    response.raise_for_status()
    return response.json()


def resolve_surveys(survey_codes=None):
    survey_codes = set(survey_codes or {
        INJECTION_SURVEY_CODE,
        TECHNICIAN_SURVEY_CODE,
        TAHER_INJECTION_SURVEY_CODE,
        TAHER_SUTURE_SURVEY_CODE,
    })
    numeric_ids = {code: int(code) for code in survey_codes if str(code).isdigit()}
    survey_codes -= set(numeric_ids)
    cached = {code: get_state(f"survey_id:{code}") for code in survey_codes}
    if all(cached.values()):
        return {**numeric_ids, **{code: int(survey_id) for code, survey_id in cached.items()}}

    folders = porsline_get("/api/folders/")
    found = {}
    for folder in folders:
        for survey in folder.get("surveys", []):
            candidates = {
                str(survey.get("preview_code") or ""),
                str(survey.get("url_slug") or ""),
                str(survey.get("report_code") or ""),
            }
            for code in survey_codes:
                if code in candidates:
                    found[code] = int(survey["id"])

    missing = survey_codes - set(found)
    if missing:
        raise RuntimeError("Could not resolve survey code(s): " + ", ".join(sorted(missing)))
    for code, survey_id in found.items():
        set_state(f"survey_id:{code}", survey_id)
    return {**numeric_ids, **found}


def fetch_results(survey_id):
    first = porsline_get(
        f"/api/v2/surveys/{survey_id}/responses/results-table/",
        params={"page": 1, "page_size": 1000},
    )
    headers = first.get("header", [])
    rows = list(first.get("body", []))
    total = int(first.get("responders_count", len(rows)))

    page = 2
    while len(rows) < total:
        batch = porsline_get(
            f"/api/v2/surveys/{survey_id}/responses/results-table/",
            params={"page": page, "page_size": 1000},
        ).get("body", [])
        if not batch:
            break
        rows.extend(batch)
        page += 1
    return headers, rows, total


def clean_header(value):
    return re.sub(r"[\s‌:\-_()（）]+", "", str(value or "")).lower()


def header_label(header, index):
    if isinstance(header, str):
        return header
    if isinstance(header, dict):
        for key in (
            "title", "name", "text", "label", "question_title", "question_text",
            "alt_name", "display_name", "header",
        ):
            value = header.get(key)
            if isinstance(value, str) and value.strip():
                return value
        for value in header.values():
            if isinstance(value, dict):
                nested = header_label(value, index)
                if nested != f"column_{index}":
                    return nested
    return f"column_{index}"


def scalar_value(value):
    if isinstance(value, list):
        if len(value) == 1:
            return scalar_value(value[0])
        return "، ".join(str(scalar_value(item)) for item in value if item not in (None, ""))
    if not isinstance(value, dict):
        return value
    for key in ("url", "file_url", "download_url", "value", "answer", "text", "name", "display_value", "response", "result"):
        candidate = value.get(key)
        if candidate not in (None, ""):
            return scalar_value(candidate)
    return value


def row_to_mapping(headers, row):
    labels = [header_label(header, i) for i, header in enumerate(headers)]
    if isinstance(row, list):
        return {labels[i]: scalar_value(value) for i, value in enumerate(row) if i < len(labels)}
    if isinstance(row, dict):
        result = {str(key): scalar_value(value) for key, value in row.items()}
        nested_values = (
            row.get("values") or row.get("answers") or row.get("cells")
            or row.get("data") or row.get("row") or row.get("responses")
        )
        if isinstance(nested_values, list):
            result.update({labels[i]: scalar_value(value) for i, value in enumerate(nested_values) if i < len(labels)})
            for cell in nested_values:
                if not isinstance(cell, dict):
                    continue
                cell_label = header_label(cell, -1)
                cell_value = scalar_value(cell)
                if cell_label != "column_-1" and cell_value is not cell:
                    result[cell_label] = cell_value
        elif isinstance(nested_values, dict):
            nested_items = list(nested_values.items())
            result.update({labels[i]: scalar_value(item[1]) for i, item in enumerate(nested_items) if i < len(labels)})
            result.update({str(key): scalar_value(value) for key, value in nested_items})

        # Some API versions wrap the positional row in an undocumented list field.
        if not any(clean_header(label) in {clean_header(key) for key in result} for label in labels):
            for value in row.values():
                if isinstance(value, list) and len(value) >= len(labels):
                    result.update({labels[i]: scalar_value(value[i]) for i in range(len(labels))})
                    break
        for i, header in enumerate(headers):
            if not isinstance(header, dict):
                continue
            candidates = []
            for key in ("id", "key", "column_id", "object_id", "question_id"):
                if header.get(key) is not None:
                    candidates.extend([header[key], str(header[key])])
            for candidate in candidates:
                if candidate in row:
                    result[labels[i]] = scalar_value(row[candidate])
                    break
        return result
    raise ValueError(f"Unsupported Porsline row type: {type(row).__name__}")


def find_value(mapping, candidates):
    normalized = {clean_header(k): v for k, v in mapping.items()}
    for candidate in candidates:
        needle = clean_header(candidate)
        if needle in normalized and normalized[needle] not in (None, ""):
            return str(normalized[needle]).strip()
    for candidate in candidates:
        needle = clean_header(candidate)
        for key, value in normalized.items():
            if needle and needle in key and value not in (None, ""):
                return str(value).strip()
    return ""


def response_key(mapping):
    identifier = find_value(
        mapping,
        ["شناسه پاسخ دهنده", "شناسه پاسخ‌دهنده", "responder id", "response id", "id"],
    )
    if identifier:
        return identifier
    raw = json.dumps(mapping, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def extract_person(mapping):
    persian_name = find_value(
        mapping,
        ["نام و نام‌خانوادگی(فارسی)", "نام و نام خانوادگی فارسی", "نام فارسی"],
    )
    english_name = find_value(
        mapping,
        ["نام و نام خانوادگی(انگلیسی)", "نام و نام خانوادگی انگلیسی", "نام انگلیسی"],
    )
    national_id = find_value(mapping, ["کد ملی", "کدملی", "national id", "national code"])
    national_id = re.sub(r"\D", "", national_id)
    return {
        "persian_name": persian_name,
        "english_name": english_name,
        "national_id": national_id,
    }


def extract_certificate_person(mapping):
    person = extract_person(mapping)
    gender = find_value(mapping, ["جنسیت", "gender"])
    photo_candidates = {
        clean_header("عکس خود را به صورت اسکن شده بارگزاری کنید."),
        clean_header("عکس خود را به صورت اسکن شده بارگذاری کنید."),
    }
    photo_value = next(
        (value for key, value in mapping.items()
         if any(candidate == clean_header(key) or candidate in clean_header(key) for candidate in photo_candidates)),
        "",
    )
    serialized = json.dumps(photo_value, ensure_ascii=False, default=str) if isinstance(photo_value, (dict, list)) else str(photo_value)
    url_match = re.search(r"https?://[^\s'\"،\\]+", serialized)
    person.update({"gender": gender, "photo_url": url_match.group(0) if url_match else ""})
    return person


def certificate_honorific(gender):
    value = str(gender or "").strip().casefold()
    if any(word in value for word in ("خانم", "زن", "female")):
        return "خانم"
    if any(word in value for word in ("آقا", "مرد", "male")):
        return "آقای"
    return ""


def safe_filename(value):
    cleaned = re.sub(r"[\\/:*?\"<>|\r\n]+", "-", str(value or "")).strip(" .-")
    return cleaned[:100] or "مدارک"


def download_portrait(url, target_path):
    if not str(url or "").startswith(("http://", "https://")):
        raise ValueError("آدرس عکس معتبر نیست")
    source_host = urlparse(str(url)).hostname
    porsline_host = urlparse(PORSLINE_BASE_URL).hostname
    headers = {"Authorization": f"API-Key {PORSLINE_API_KEY}"} if source_host == porsline_host else {}
    response = requests.get(url, headers=headers, timeout=90)
    response.raise_for_status()
    with Image.open(io.BytesIO(response.content)) as source:
        image = ImageOps.exif_transpose(source).convert("RGB")
        # نسبت دقیق قاب عکس در فایل نمونه؛ تصویر کامل می‌ماند و فقط حاشیه سفید می‌گیرد.
        frame_ratio = 120.64 / 149.85
        if image.width / image.height > frame_ratio:
            canvas_size = (image.width, max(1, round(image.width / frame_ratio)))
        else:
            canvas_size = (max(1, round(image.height * frame_ratio)), image.height)
        canvas = Image.new("RGB", canvas_size, "white")
        canvas.paste(image, ((canvas.width - image.width) // 2, (canvas.height - image.height) // 2))
        canvas.save(target_path, "PNG", optimize=True)


def build_certificate_powerpoint(students, course_title):
    if not CERTIFICATE_TEMPLATE_PATH.exists():
        raise RuntimeError("قالب پاورپوینت مدارک در برنامه وجود ندارد.")
    temp_dir = tempfile.TemporaryDirectory(prefix="porsline-certificates-")
    root = Path(temp_dir.name)
    prepared = []
    successful = []
    failed = []
    try:
        for index, student in enumerate(students, start=1):
            portrait_path = root / f"portrait-{index}.png"
            try:
                download_portrait(student["photo_url"], portrait_path)
            except Exception as exc:
                log.warning("Could not download portrait for %s: %s", student["persian_name"], exc)
                failed.append({"name": student["persian_name"], "reason": "عکس قابل دریافت یا خواندن نیست"})
                continue
            prepared.append({**student, "portrait_path": str(portrait_path)})
            successful.append(student)
        if not prepared:
            return None, None, [], failed
        dated = datetime.now(ZoneInfo(DISPLAY_TIMEZONE)).strftime("%Y-%m-%d")
        filename = f"مدارک - {safe_filename(course_title)} - {dated} - {len(prepared)} نفر.pptx"
        output_path = root / filename
        config_path = root / "config.json"
        config_path.write_text(json.dumps({
            "template_path": str(CERTIFICATE_TEMPLATE_PATH),
            "output_path": str(output_path),
            "students": prepared,
        }, ensure_ascii=False), encoding="utf-8")
        process = subprocess.run(
            [sys.executable, str(CERTIFICATE_WORKER_PATH), str(config_path)],
            capture_output=True, text=True, timeout=300,
        )
        if process.returncode != 0 or not output_path.exists():
            log.error("Certificate worker failed: %s", process.stderr[-3000:])
            raise RuntimeError("ساخت فایل PowerPoint ناموفق بود.")
        stream = io.BytesIO(output_path.read_bytes())
        stream.seek(0)
        return filename, stream, successful, failed
    finally:
        temp_dir.cleanup()


def is_processed(survey_code, key):
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM processed_responses WHERE survey_code=%s AND response_key=%s",
            (survey_code, key),
        )
        return cur.fetchone() is not None


def is_certificate_generated(survey_code, key):
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM self_certificate_items WHERE survey_code=%s AND response_key=%s",
            (survey_code, key),
        )
        return cur.fetchone() is not None


def record_certificate_batch(students, report_names, output_mode):
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """INSERT INTO self_certificate_batches(report_names, row_count, output_mode)
               VALUES(%s, %s, %s) RETURNING id, created_at""",
            (json.dumps(report_names, ensure_ascii=False), len(students), output_mode),
        )
        batch_id, created_at = cur.fetchone()
        jy, jm, jd = gregorian_to_jalali(created_at.year, created_at.month, created_at.day)
        short_id = f"P-{jy % 100:02d}{jm:02d}{jd:02d}-{batch_id}"
        cur.execute("UPDATE self_certificate_batches SET short_id=%s WHERE id=%s", (short_id, batch_id))
        cur.executemany(
            """INSERT INTO self_certificate_items(
                batch_id, survey_code, response_key, report_name, teacher_name, persian_name, national_id
            ) VALUES(%s,%s,%s,%s,%s,%s,%s)""",
            [(batch_id, row["survey_code"], row["response_key"], row["report_name"],
              row["teacher_name"], row["persian_name"], row["national_id"]) for row in students],
        )
        conn.commit()
    return short_id


def collect_certificate_rows(code, survey_id, settings):
    headers, rows, total = fetch_results(survey_id)
    students, skipped = [], []
    for raw_row in rows:
        mapping = row_to_mapping(headers, raw_row)
        key = response_key(mapping)
        if is_certificate_generated(code, key):
            continue
        person = extract_certificate_person(mapping)
        honorific = certificate_honorific(person.get("gender"))
        missing = []
        if not person.get("persian_name"):
            missing.append("نام و نام خانوادگی فارسی")
        if not person.get("national_id"):
            missing.append("کد ملی")
        if not honorific:
            missing.append("جنسیت")
        if not person.get("photo_url"):
            missing.append("عکس")
        if missing:
            skipped.append({"name": person.get("persian_name") or f"پاسخ {key[:8]}", "reason": "، ".join(missing)})
            continue
        students.append({
            **person, "honorific": honorific, "response_key": key, "survey_code": code,
            "report_name": settings["report_name"], "teacher_name": settings["teacher_name"],
            "course_title": settings["course_title"], "duration": settings["course_duration"],
            "instructor": settings["certificate_instructor"], "venue": settings["certificate_venue"],
            "organization": settings["certificate_organization"],
        })
    return students, skipped, total


def send_skipped_certificates(skipped, chat_id=None):
    if not skipped:
        return
    lines = ["⚠️ برای افراد زیر مدرک ساخته نشد:"]
    lines.extend(f"• {item['name']}: {item['reason']}" for item in skipped)
    send_message("\n".join(lines)[:4000], chat_id=chat_id)


def run_certificate_export(survey_codes, output_mode="combined", chat_id=None):
    if not RUN_LOCK.acquire(blocking=False):
        return {"status": "already-running"}
    try:
        settings_list = [get_form_settings(code) for code in survey_codes]
        if any(not item or item["certificate_type"] != "self" for item in settings_list):
            raise ValueError("⚠️ فقط فرم‌هایی را انتخاب کنید که مدرکشان را خودتان صادر می‌کنید.")
        ids = resolve_surveys(set(survey_codes))
        grouped, all_skipped = [], []
        for settings in settings_list:
            students, skipped, _total = collect_certificate_rows(
                settings["survey_code"], ids[settings["survey_code"]], settings
            )
            grouped.append((settings, students))
            all_skipped.extend(skipped)
        sent = []
        groups = grouped if output_mode == "separate" else [(
            {"course_title": " + ".join(dict.fromkeys(item[0]["course_title"] for item in grouped)),
             "report_name": "چند فرم"},
            [student for _settings, students in grouped for student in students],
        )]
        for settings, students in groups:
            if not students:
                continue
            filename, stream, built_students, failed = build_certificate_powerpoint(
                students, settings["course_title"]
            )
            all_skipped.extend(failed)
            if not built_students:
                continue
            caption = (
                f"🎓 مدارک آماده چاپ\nتعداد: {len(built_students)} نفر\n"
                f"فرم‌ها: {'، '.join(dict.fromkeys(row['report_name'] for row in built_students))}\n"
                "وضعیت: 🖨 ارسال‌شده به چاپخانه"
            )
            send_document(
                filename, stream, caption, chat_id,
                "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            )
            record_certificate_batch(
                built_students, list(dict.fromkeys(row["report_name"] for row in built_students)), output_mode
            )
            mark_processed([(row["survey_code"], row["response_key"]) for row in built_students])
            sent.append(filename)
        send_skipped_certificates(all_skipped, chat_id)
        if not sent:
            send_message("ℹ️ ثبت جدیدِ کامل و قابل ساختی برای مدرک وجود نداشت.", chat_id=chat_id)
        return {"status": "sent" if sent else "no-new-rows", "files": sent, "skipped": all_skipped}
    finally:
        RUN_LOCK.release()


def mark_processed(items):
    with db() as conn, conn.cursor() as cur:
        cur.executemany(
            """
            INSERT INTO processed_responses(survey_code, response_key)
            VALUES(%s, %s) ON CONFLICT DO NOTHING
            """,
            items,
        )
        cur.executemany(
            "INSERT INTO response_send_history(survey_code, response_key) VALUES(%s, %s)",
            items,
        )
        conn.commit()


def national_id_formula(row_number):
    b = f'SUBSTITUTE($B{row_number},"/","")'
    return (
        f'=IFERROR(IF(AND(LEN({b})=10,AND(LEFT({b},10)<>REPT(ROW($1:$9),10)),'
        f'OR(AND(MOD(SUM(MID({b},ROW($1:$9),1)*(11-ROW($1:$9))),11)<2,'
        f'--RIGHT({b})=MOD(SUM(MID({b},ROW($1:$9),1)*(11-ROW($1:$9))),11)),'
        f'--RIGHT({b})=(11-MOD(SUM(MID({b},ROW($1:$9),1)*(11-ROW($1:$9))),11)))),'
        "TRUE,FALSE),FALSE)"
    )


def display_national_id(national_id):
    value = str(national_id or "").strip()
    return f"/{value}" if value.startswith("0") else value


def build_report(rows, total_count, report_name, colored=False):
    template_bytes = base64.b64decode(TEMPLATE_PATH.read_text(encoding="ascii"))
    workbook = load_workbook(io.BytesIO(template_bytes))
    sheet = workbook.active

    if sheet.max_column > 5:
        sheet.delete_cols(6, sheet.max_column - 5)

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    sheet.sheet_view.rightToLeft = True
    sheet.cell(1, 1, "نام فارسی")
    sheet.cell(1, 2, "کد ملی فارسی")
    sheet.cell(1, 3, "TRUE")
    sheet.cell(1, 4, "نام انگلیسی")
    sheet.cell(1, 5, "کد ملی انگلیسی")

    for row_number, person in enumerate(rows, start=2):
        national_id = display_national_id(person["national_id"])
        sheet.cell(row_number, 1, person["persian_name"])
        sheet.cell(row_number, 2, national_id)
        sheet.cell(row_number, 3, national_id_formula(row_number))
        sheet.cell(row_number, 4, person["english_name"])
        sheet.cell(row_number, 5, national_id)
        for col in range(1, 6):
            cell = sheet.cell(row_number, col)
            if col in (2, 5):
                cell.number_format = "@"
            if colored:
                color_key = person.get("color_key") or "white"
                rgb = TEACHER_COLORS.get(color_key, TEACHER_COLORS["white"])[1]
                cell.fill = PatternFill(fill_type="solid", fgColor=rgb)

    filename = f"{report_name} {total_count}.xlsx"
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return filename, stream


def report_name_range(rows):
    names = [str(row.get("persian_name") or row.get("english_name") or "").strip() for row in rows]
    names = [name for name in names if name]
    if not names:
        return "نام افراد ثبت نشده است"
    if len(names) == 1:
        return f"شامل یک نفر: {names[0]}"
    return f"از {names[0]} تا {names[-1]}"


def report_caption(rows, report_names, empty_names=None, batch_short_id=None, colored=False):
    lines = []
    if batch_short_id:
        lines.append(f"شناسه گزارش: {batch_short_id}")
    lines.extend([f"👥 {report_name_range(rows)}", f"تعداد: {len(rows)} نفر"])
    if report_names:
        lines.append("فرم‌ها: " + "، ".join(report_names))
    if empty_names:
        lines.append("بدون ثبت جدید: " + "، ".join(empty_names))
    if colored:
        seen = []
        for row in rows:
            pair = (row.get("teacher_name", ""), row.get("color_key") or "white")
            if pair not in seen:
                seen.append(pair)
        if seen:
            lines.append("راهنمای رنگ: " + " | ".join(
                f"{teacher}: {TEACHER_COLORS.get(color, TEACHER_COLORS['white'])[0]}"
                for teacher, color in seen
            ))
    caption = "\n".join(lines)
    return caption if len(caption) <= 1024 else caption[:1021] + "…"


def create_report_batch(rows, report_names, empty_names, is_combined):
    if not rows:
        return None
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO report_batches(report_names, empty_report_names, row_count, first_name, last_name, is_combined)
            VALUES(%s, %s, %s, %s, %s, %s) RETURNING id, created_at
            """,
            (
                json.dumps(report_names, ensure_ascii=False),
                json.dumps(empty_names, ensure_ascii=False),
                len(rows), rows[0].get("persian_name", ""), rows[-1].get("persian_name", ""),
                bool(is_combined),
            ),
        )
        batch_id, created_at = cur.fetchone()
        short_id = f"R-{created_at.astimezone(ZoneInfo(DISPLAY_TIMEZONE)).strftime('%y%m%d')}-{batch_id:04d}"
        cur.execute("UPDATE report_batches SET short_id=%s WHERE id=%s", (short_id, batch_id))
        cur.executemany(
            """
            INSERT INTO report_batch_items(
                batch_id, survey_code, response_key, report_name, teacher_name,
                color_key, persian_name, english_name, national_id, row_position
            ) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            [(
                batch_id, row["survey_code"], row["response_key"], row["report_name"],
                row["teacher_name"], row.get("color_key") or "white",
                row.get("persian_name", ""), row.get("english_name", ""),
                row.get("national_id", ""), position,
            ) for position, row in enumerate(rows, start=1)],
        )
        conn.commit()
    return {"id": batch_id, "short_id": short_id, "created_at": created_at}


def get_batch_items(batch_id):
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT survey_code, response_key, report_name, teacher_name, color_key,
                   persian_name, english_name, national_id
            FROM report_batch_items WHERE batch_id=%s ORDER BY row_position
            """,
            (batch_id,),
        )
        return [{
            "survey_code": row[0], "response_key": row[1], "report_name": row[2],
            "teacher_name": row[3], "color_key": row[4] or "white",
            "persian_name": row[5], "english_name": row[6], "national_id": row[7],
        } for row in cur.fetchall()]


def recent_batches(stage=None, limit=10):
    conditions = []
    if stage == "receive":
        conditions.append("nursing_received_at IS NULL")
    elif stage == "undo":
        conditions.append("nursing_received_at IS NOT NULL")
    where = " WHERE " + " AND ".join(conditions) if conditions else ""
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT id, short_id, created_at, report_names, row_count, is_combined, "
            "nursing_received_at FROM report_batches" + where +
            " ORDER BY created_at DESC LIMIT %s",
            (limit,),
        )
        return [{
            "id": row[0], "short_id": row[1], "created_at": row[2],
            "report_names": json.loads(row[3]), "row_count": row[4], "is_combined": row[5],
            "nursing_received_at": row[6],
        } for row in cur.fetchall()]


def update_batch_status(batch_id, action):
    assignments = {
        "receive": "nursing_received_at=NOW()",
        "undo_receive": "nursing_received_at=NULL, posted_at=NULL",
    }
    if action not in assignments:
        raise ValueError("⚠️ عملیات وضعیت معتبر نیست.")
    with db() as conn, conn.cursor() as cur:
        cur.execute(f"UPDATE report_batches SET {assignments[action]} WHERE id=%s", (batch_id,))
        if cur.rowcount != 1:
            raise ValueError("⚠️ گزارش موردنظر پیدا نشد.")
        conn.commit()


def get_batch(batch_id):
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, short_id, created_at, report_names, empty_report_names, row_count,
                   is_combined, nursing_received_at
            FROM report_batches WHERE id=%s
            """,
            (batch_id,),
        )
        row = cur.fetchone()
    if not row:
        return None
    return {
        "id": row[0], "short_id": row[1], "created_at": row[2],
        "report_names": json.loads(row[3]), "empty_names": json.loads(row[4]),
        "row_count": row[5], "is_combined": row[6], "nursing_received_at": row[7],
    }


def delete_report_batch(batch_id):
    with db() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM report_batches WHERE id=%s", (batch_id,))
        conn.commit()


def owner_chat_id():
    chat_id = get_state("bot_owner_chat_id")
    if not chat_id:
        raise RuntimeError("Bot owner has not been registered")
    return str(chat_id)


def send_document(filename, stream, caption, chat_id=None, mime_type=None):
    response = requests.post(
        f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument",
        data={"chat_id": chat_id or owner_chat_id(), "caption": caption},
        files={"document": (filename, stream, mime_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=120,
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("ok"):
        raise RuntimeError(f"Telegram rejected document: {payload}")


def send_message(text, chat_id=None, reply_markup=None):
    payload = {"chat_id": chat_id or owner_chat_id(), "text": text}
    if reply_markup is not None:
        payload["reply_markup"] = reply_markup
    response = requests.post(
        f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
        json=payload,
        timeout=45,
    )
    response.raise_for_status()


def is_group_admin(user_id):
    response = requests.get(
        f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getChatMember",
        params={"chat_id": TELEGRAM_CHAT_ID, "user_id": user_id},
        timeout=45,
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("ok"):
        return False
    return (payload.get("result") or {}).get("status") in {"creator", "administrator"}


def parse_add_form(text):
    match = re.match(
        r"^/add_form(?:@\w+)?\s+(/[a-z][a-z0-9_]{1,30})\s+"
        r"(https?://survey\.porsline\.ir/s/([A-Za-z0-9]+))\s*\|\s*(.+?)\s*$",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        raise ValueError(
            "فرمت دستور درست نیست. نمونه:\n"
            "/add_form /my_form https://survey.porsline.ir/s/AbC123 | نام فایل"
        )
    command = match.group(1).lower()
    survey_code = match.group(3)
    report_name = match.group(4).strip()
    if command in UTILITY_COMMANDS or command in SINGLE_REPORT_COMMANDS:
        raise ValueError("این دستور رزرو شده است؛ یک دستور متفاوت انتخاب کنید.")
    if len(report_name) > 80:
        raise ValueError("نام فایل باید حداکثر ۸۰ کاراکتر باشد.")
    return command, survey_code, report_name


def parse_survey_identifier(text):
    value = str(text or "").strip()
    link_match = re.fullmatch(
        r"https?://(?:survey\.)?porsline\.ir/(?:s|survey)/([A-Za-z0-9_-]+)(?:[/?#].*)?",
        value,
        flags=re.IGNORECASE,
    )
    if link_match:
        return link_match.group(1)
    if re.fullmatch(r"[A-Za-z0-9_-]{3,100}", value):
        return value
    raise ValueError("شناسه معتبر نیست. کد فرم، شناسه عددی یا لینک کامل فرم پرس‌لاین را بفرستید.")


def validate_survey_access(survey_code):
    survey_id = resolve_surveys({survey_code})[survey_code]
    porsline_get(
        f"/api/v2/surveys/{survey_id}/responses/results-table/",
        params={"page": 1, "page_size": 1},
    )
    return survey_id


def collect_new_rows(code, survey_id, include_processed=False):
    headers, rows, total = fetch_results(survey_id)
    log.info("Survey %s headers: %s", code, [header_label(h, i) for i, h in enumerate(headers)])
    if rows:
        first = rows[0]
        shape = {
            "row_type": type(first).__name__,
            "keys": list(first.keys()) if isinstance(first, dict) else None,
            "value_types": {str(k): type(v).__name__ for k, v in first.items()} if isinstance(first, dict) else None,
            "length": len(first) if isinstance(first, (dict, list)) else None,
        }
        log.info("Survey %s row shape: %s", code, shape)
    new_people = []
    keys = []
    for raw_row in rows:
        mapping = row_to_mapping(headers, raw_row)
        key = response_key(mapping)
        if not include_processed and is_processed(code, key):
            continue
        person = extract_person(mapping)
        required_fields = ("persian_name", "national_id")
        if not all(person[name] for name in required_fields):
            missing = [name for name in required_fields if not person[name]]
            log.warning("Skipped incomplete response %s from survey %s; missing=%s", key, code, missing)
            continue
        new_people.append(person)
        keys.append((code, key))
    return new_people, keys, total


def run_report(include_processed=False, all_forms=False, selected_reports=None, combine=None):
    if not RUN_LOCK.acquire(blocking=False):
        return {"status": "already-running"}
    try:
        require_settings()
        init_db()
        reports = selected_reports or [
            (INJECTION_SURVEY_CODE, "تزریقات"),
            (TECHNICIAN_SURVEY_CODE, "تکنسین داروخانه"),
        ]
        if all_forms:
            reports = all_report_definitions("nursing")
        ids = resolve_surveys({code for code, _ in reports})
        sent_files = []
        results = {}
        form_meta = {code: (name, teacher) for _source, code, name, teacher in get_form_records(active_only=False)}
        colors = get_teacher_colors()
        collected = []
        empty_names = []
        for code, report_name in reports:
            people, keys, total = collect_new_rows(code, ids[code], include_processed)
            results[code] = {"report": report_name, "rows": len(people), "total": total}
            if not people:
                empty_names.append(report_name)
                continue
            teacher_name = form_meta.get(code, (report_name, "نامشخص"))[1]
            for person, (_key_code, response_id) in zip(people, keys):
                collected.append({
                    **person, "survey_code": code, "response_key": response_id,
                    "report_name": report_name, "teacher_name": teacher_name,
                    "color_key": colors.get(teacher_name) or "white", "survey_total": total,
                })

        combine = (len(reports) > 1 and not include_processed) if combine is None else bool(combine)
        if collected and combine:
            missing_colors = sorted({row["teacher_name"] for row in collected if not colors.get(row["teacher_name"])})
            if missing_colors:
                raise ValueError("⚠️ ابتدا از بخش مدیریت برای این مدرس‌ها رنگ تعیین کنید: " + "، ".join(missing_colors))
            report_names = [name for _code, name in reports if name not in empty_names]
            batch = create_report_batch(collected, report_names, empty_names, True)
            dated = batch["created_at"].astimezone(ZoneInfo(DISPLAY_TIMEZONE)).strftime("%Y-%m-%d")
            white_name, white_stream = build_report(
                collected, len(collected), f"گزارش نظام پرستاری - {dated}", colored=False
            )
            color_name, color_stream = build_report(
                collected, len(collected), f"گزارش ترکیبی رنگی - {dated}", colored=True
            )
            common_caption = report_caption(collected, report_names, empty_names, batch["short_id"])
            try:
                send_document(white_name, white_stream, common_caption)
                send_document(
                    color_name, color_stream,
                    report_caption(collected, report_names, empty_names, batch["short_id"], colored=True),
                )
            except Exception:
                delete_report_batch(batch["id"])
                raise
            mark_processed([(row["survey_code"], row["response_key"]) for row in collected])
            sent_files.extend([white_name, color_name])
        elif collected:
            for code, report_name in reports:
                form_rows = [row for row in collected if row["survey_code"] == code]
                if not form_rows:
                    continue
                total = results[code]["total"]
                filename, stream = build_report(form_rows, total, report_name, colored=False)
                if include_processed:
                    caption = report_caption(form_rows, [report_name])
                else:
                    batch = create_report_batch(form_rows, [report_name], [], False)
                    caption = report_caption(form_rows, [report_name], batch_short_id=batch["short_id"])
                try:
                    send_document(filename, stream, caption)
                except Exception:
                    if not include_processed:
                        delete_report_batch(batch["id"])
                    raise
                if not include_processed:
                    mark_processed([(row["survey_code"], row["response_key"]) for row in form_rows])
                sent_files.append(filename)

        if not sent_files:
            if include_processed:
                send_message("در فرم‌ها پاسخی برای ارسال وجود نداشت.")
            else:
                detail = "\nفرم‌های بدون ثبت جدید: " + "، ".join(empty_names) if empty_names else ""
                send_message("در این دوره پاسخ جدیدی برای ارسال وجود نداشت." + detail)
        return {
            "status": "sent" if sent_files else "no-new-rows",
            "files": sent_files,
            "reports": results,
            "empty_reports": empty_names,
        }
    finally:
        RUN_LOCK.release()


def run_single_report(command):
    report = get_report_for_command(command)
    if not report:
        return {"status": "unknown-command"}
    return run_report(selected_reports=[report], combine=False)


def run_status():
    if not RUN_LOCK.acquire(blocking=False):
        return None
    try:
        require_settings()
        init_db()
        records = get_form_records()
        reports = [(code, name) for _source, code, name, _teacher in records]
        ids = resolve_surveys({code for code, _ in reports})
        status_rows = []
        for code, report_name in reports:
            settings = get_form_settings(code) or {"certificate_type": "nursing"}
            if settings["certificate_type"] == "self":
                people, _skipped, total = collect_certificate_rows(code, ids[code], settings)
            else:
                people, _keys, total = collect_new_rows(code, ids[code])
            status_rows.append((report_name, len(people), total))
        return status_rows
    finally:
        RUN_LOCK.release()


def help_text():
    return (
        "راهنمای دستورات ربات:\n\n"
        "/zanyar_t — پاسخ‌های جدید تزریقات\n"
        "/zanyar_tek — پاسخ‌های جدید تکنسین داروخانه\n"
        "/taher_t — پاسخ‌های جدید تزریقات خانم طاهرخانی\n"
        "/taher_b — پاسخ‌های جدید بخیه خانم طاهرخانی\n"
        "/report_all_new — پاسخ‌های جدید همه فرم‌ها\n"
        "/report_all — همه پاسخ‌های همه فرم‌ها (نیازمند تأیید)\n"
        "/status — تعداد کل و جدید همه فرم‌ها\n"
        "/forms — فهرست فرم‌ها و دستورهایشان\n"
        "/add_form — افزودن فرم جدید (فقط مدیر گروه)\n"
        "/remove_form — غیرفعال‌کردن فرم افزوده‌شده (فقط مدیر گروه)\n"
        "/help — نمایش همین راهنما"
    )


def friendly_error_message(exc):
    if isinstance(exc, ValueError):
        return str(exc)
    if isinstance(exc, (requests.Timeout, requests.ConnectionError)):
        return "ارتباط اینترنتی با پرسلاین یا تلگرام برقرار نشد. چند دقیقه بعد دوباره تلاش کنید."
    if isinstance(exc, requests.HTTPError):
        response = getattr(exc, "response", None)
        status_code = getattr(response, "status_code", None)
        if status_code in (401, 403):
            return "دسترسی به یکی از سرویس‌ها رد شد. لطفاً کلید API و توکن ربات را بررسی کنید."
        return "یکی از سرویس‌های پرسلاین یا تلگرام موقتاً پاسخ نداد. کمی بعد دوباره تلاش کنید."
    if isinstance(exc, psycopg.Error):
        return "ارتباط با پایگاه‌داده برقرار نشد. کمی بعد دوباره تلاش کنید."
    return "هنگام آماده‌سازی گزارش خطایی رخ داد. لطفاً چند دقیقه بعد دوباره تلاش کنید."


def main_menu():
    teachers = get_teachers()
    keyboard = []
    for index in range(0, len(teachers), 2):
        keyboard.append([{"text": f"👨‍🏫 {name}"} for name in teachers[index:index + 2]])
    keyboard.extend([
        [{"text": "📊 وضعیت فرم‌ها"}, {"text": "📦 گزارش همه فرم‌ها"}],
        [{"text": "☑️ انتخاب چند فرم"}, {"text": "🎓 مدارک نظام پرستاری"}],
        [{"text": "🔍 جست‌وجوی ثبت‌نام‌کنندگان"}, {"text": "🩺 بررسی سلامت"}],
        [{"text": "⚙️ مدیریت ربات"}],
        [{"text": "✖️ بستن منو"}],
    ])
    return {"keyboard": keyboard, "resize_keyboard": True, "is_persistent": True}


def teacher_menu(teacher_name):
    keyboard = []
    for _source, survey_code, report_name, _teacher in get_form_records(teacher_name=teacher_name):
        keyboard.append([{"text": custom_form_button_label(report_name, survey_code)}])
    keyboard.extend([
        [{"text": "✏️ ویرایش نام مدرس"}],
        [{"text": "🗑 غیرفعال‌کردن مدرس و تمام فرم‌ها"}],
    ])
    keyboard.append([{"text": "🔙 بازگشت به منوی اصلی"}])
    return {"keyboard": keyboard, "resize_keyboard": True, "is_persistent": True}


def teacher_selection_menu():
    keyboard = [[{"text": f"👨‍🏫 {name}"}] for name in get_teachers()]
    keyboard.extend([
        [{"text": "➕ مدرس جدید"}],
        [{"text": "❌ لغو ثبت فرم"}],
        [{"text": "❌ انصراف"}],
    ])
    return {"keyboard": keyboard, "resize_keyboard": True, "one_time_keyboard": True}


CERTIFICATE_TYPE_MENU = {
    "keyboard": [
        [{"text": "🏥 برای نظام پرستاری"}],
        [{"text": "🎓 خودم مدرک صادر می‌کنم"}],
        [{"text": "❌ انصراف"}],
    ],
    "resize_keyboard": True,
    "one_time_keyboard": True,
}


MANAGEMENT_MENU = {
    "keyboard": [
        [{"text": "➕ ثبت فرم جدید"}],
        [{"text": "➕ افزودن مدرس"}, {"text": "🎨 رنگ مدرس‌ها"}],
        [{"text": "🔄 انتقال فرم بین مدرس‌ها"}],
        [{"text": "♻️ بازیابی موارد غیرفعال"}],
        [{"text": "🔙 بازگشت به منوی اصلی"}],
    ],
    "resize_keyboard": True,
    "is_persistent": True,
}

NURSING_MENU = {
    "keyboard": [
        [{"text": "1️⃣ مدارک از نظام رسید"}],
        [{"text": "📊 وضعیت گزارش‌های اخیر"}],
        [{"text": "↩️ اصلاح انتخاب اشتباه"}],
        [{"text": "📘 راهنمای این بخش"}],
        [{"text": "🔙 بازگشت به منوی اصلی"}],
        [{"text": "❌ انصراف"}],
    ],
    "resize_keyboard": True,
    "is_persistent": True,
}


def nursing_help_text():
    return (
        "🎓 راهنمای ساده مدارک نظام پرستاری\n\n"
        "هر بار که از ثبت‌نام‌های جدید فایل می‌گیرید، ربات آن را با یک شناسه مثل R-140505-12 ثبت می‌کند. "
        "از همان لحظه، وضعیت تمام افراد آن فایل «منتظر دریافت از نظام» است.\n\n"
        "1️⃣ وقتی مدارک همه افراد یک گزارش از نظام پرستاری به دست شما رسید:\n"
        "دکمه «مدارک از نظام رسید» را بزنید و همان گزارش را انتخاب کنید.\n\n"
        "🔎 بعد از این کار، در جست‌وجوی هر فرد، مرحله و تاریخ مدرک او نمایش داده می‌شود.\n\n"
        "↩️ اگر گزارش اشتباهی را انتخاب کردید، از «اصلاح انتخاب اشتباه» استفاده کنید."
    )


def nursing_status_text(limit=10):
    batches = recent_batches(None, limit)
    if not batches:
        return "📊 هنوز هیچ خروجی جدیدی برای پیگیری مدارک ثبت نشده است."
    lines = ["📊 وضعیت گزارش‌های اخیر:"]
    for batch in batches:
        if batch["nursing_received_at"]:
            status = "✅ از نظام پرستاری دریافت شده"
        else:
            status = "📤 به نظام پرستاری ارسال شده؛ هنوز دریافت نشده"
        names = "، ".join(batch["report_names"])
        lines.extend([
            "",
            f"{batch['short_id']} — {batch['row_count']} نفر",
            f"📋 {names}",
            f"🗓 {format_sent_at(batch['created_at'])}",
            f"وضعیت: {status}",
        ])
    return "\n".join(lines)


def color_menu(exclude_teacher=None):
    colors = available_teacher_colors(exclude_teacher)
    keyboard = []
    for index in range(0, len(colors), 2):
        keyboard.append([{"text": f"🎨 {label}"} for _key, label, _rgb in colors[index:index + 2]])
    keyboard.append([{"text": "❌ انصراف"}])
    return {"keyboard": keyboard, "resize_keyboard": True, "one_time_keyboard": True}


def teacher_color_selection_menu():
    keyboard = [[{"text": f"🎨 مدرس • {name}"}] for name in get_teachers()]
    keyboard.extend([[{"text": "🔙 بازگشت به مدیریت"}], [{"text": "❌ انصراف"}]])
    return {"keyboard": keyboard, "resize_keyboard": True, "is_persistent": True}


def multi_form_menu(selected_codes=None, counts=None):
    selected_codes = set(selected_codes or [])
    counts = counts or {}
    keyboard = []
    for _source, code, name, _teacher in get_form_records():
        digest = custom_form_command(code).removeprefix("/form_")
        count = counts.get(code)
        mark = "✅" if code in selected_codes else "⬜"
        count_text = f" — {count} جدید" if count is not None else ""
        callback = f"multi:zero:{digest}" if count == 0 else f"multi:toggle:{digest}"
        keyboard.append([{"text": f"{mark} {name[:32]}{count_text}", "callback_data": callback}])
    keyboard.extend([
        [
            {"text": "✅ انتخاب همه", "callback_data": "multi:all"},
            {"text": "🧹 پاک‌کردن", "callback_data": "multi:clear"},
        ],
        [
            {"text": "ادامه", "callback_data": "multi:continue"},
            {"text": "❌ انصراف", "callback_data": "cancel"},
        ],
    ])
    return {"inline_keyboard": keyboard}


def batch_button_label(batch):
    names = "، ".join(batch["report_names"])
    if batch.get("nursing_received_at"):
        icon = "✅"
    else:
        icon = "⏳"
    return f"{icon} {batch['short_id']} • {batch['row_count']} نفر • {names[:30]}"


def batch_selection_menu(stage):
    batches = recent_batches(stage if stage in {"receive", "post", "undo"} else None)
    keyboard = [[{"text": batch_button_label(batch)}] for batch in batches]
    keyboard.extend([[{"text": "🔙 بازگشت به مدارک"}], [{"text": "❌ انصراف"}]])
    return batches, {"keyboard": keyboard, "resize_keyboard": True, "is_persistent": True}

SEARCH_MENU = {
    "keyboard": [[{"text": "❌ لغو جست‌وجو"}], [{"text": "❌ انصراف"}]],
    "resize_keyboard": True,
    "one_time_keyboard": True,
}


def archive_menu():
    keyboard = []
    for teacher_name in get_inactive_teachers():
        keyboard.append([{"text": f"♻️ مدرس • {teacher_name}"}])
    for _source, survey_code, report_name, _teacher in get_inactive_form_records():
        keyboard.append([{"text": f"♻️ فرم • {report_name[:42]} • {str(survey_code)[-8:]}"}])
    keyboard.append([{"text": "🔙 بازگشت به مدیریت"}])
    return {"keyboard": keyboard, "resize_keyboard": True, "is_persistent": True}


def transfer_form_menu():
    keyboard = [
        [{"text": f"🔄 {report_name[:46]} • {str(survey_code)[-8:]}"}]
        for _source, survey_code, report_name, _teacher in get_form_records()
    ]
    keyboard.append([{"text": "🔙 بازگشت به مدیریت"}])
    return {"keyboard": keyboard, "resize_keyboard": True, "is_persistent": True}


def transfer_teacher_menu(current_teacher):
    keyboard = [
        [{"text": f"👨‍🏫 {name}"}] for name in get_teachers() if name != current_teacher
    ]
    keyboard.append([{"text": "❌ لغو انتقال"}])
    keyboard.append([{"text": "❌ انصراف"}])
    return {"keyboard": keyboard, "resize_keyboard": True, "one_time_keyboard": True}

MY_MENU = {
    "keyboard": [
        [{"text": "💉 ثبتی‌های جدید تزریقات"}],
        [{"text": "💊 ثبتی‌های جدید تکنسین"}],
        [{"text": "🆕 گزارش جدید دو دوره اسکالپل"}],
        [{"text": "🔙 بازگشت به منوی اصلی"}],
    ],
    "resize_keyboard": True,
    "is_persistent": True,
}

TAHER_MENU = {
    "keyboard": [
        [{"text": "💉 ثبتی‌های جدید تزریقات طاهرخانی"}],
        [{"text": "🩹 ثبتی‌های جدید بخیه طاهرخانی"}],
        [{"text": "🆕 گزارش جدید دو دوره طاهرخانی"}],
        [{"text": "🔙 بازگشت به منوی اصلی"}],
    ],
    "resize_keyboard": True,
    "is_persistent": True,
}

EXPORT_ACTIONS = {
    "mine_injection": ("ثبتی‌های جدید تزریقات", [FIVE_REPORTS[0]], False),
    "mine_technician": ("ثبتی‌های جدید تکنسین", [FIVE_REPORTS[1]], False),
    "mine_both": ("گزارش جدید دو دوره اسکالپل", FIVE_REPORTS[:2], False),
    "taher_injection": ("ثبتی‌های جدید تزریقات طاهرخانی", [FIVE_REPORTS[2]], False),
    "taher_suture": ("ثبتی‌های جدید بخیه طاهرخانی", [FIVE_REPORTS[3]], False),
    "taher_both": ("گزارش جدید دو دوره طاهرخانی", FIVE_REPORTS[2:4], False),
    "zohiri": ("ثبتی‌های جدید مدارک خانم ظهیری", [FIVE_REPORTS[4]], False),
    "all_full": ("گزارش کامل همه فرم‌ها", None, True),
}

BUTTON_ACTIONS = {
    "💉 ثبتی‌های جدید تزریقات": "mine_injection",
    "💊 ثبتی‌های جدید تکنسین": "mine_technician",
    "🆕 گزارش جدید دو دوره اسکالپل": "mine_both",
    "💉 ثبتی‌های جدید تزریقات طاهرخانی": "taher_injection",
    "🩹 ثبتی‌های جدید بخیه طاهرخانی": "taher_suture",
    "🆕 گزارش جدید دو دوره طاهرخانی": "taher_both",
    "📄 ثبتی‌های جدید مدارک خانم ظهیری": "zohiri",
    "📦 گزارش همه فرم‌ها": "all_full",
}

CUSTOM_FORM_MENU = {
    "keyboard": [
        [{"text": "🆕 فقط ثبت‌های جدید"}],
        [{"text": "📚 همه ثبت‌ها"}],
        [{"text": "✏️ ویرایش نام فرم"}],
        [{"text": "⚙️ ویرایش مشخصات مدرک"}],
        [{"text": "🗑 غیرفعال‌کردن فرم"}],
        [{"text": "🔙 بازگشت به مدرس"}],
    ],
    "resize_keyboard": True,
    "is_persistent": True,
}


def custom_form_menu(survey_code):
    settings = get_form_settings(survey_code)
    if not settings or settings["certificate_type"] != "self":
        return CUSTOM_FORM_MENU
    return {
        "keyboard": [
            [{"text": "🎓 ساخت PowerPoint مدارک جدید"}],
            [{"text": "✏️ ویرایش نام فرم"}],
            [{"text": "⚙️ ویرایش مشخصات مدرک"}],
            [{"text": "🗑 غیرفعال‌کردن فرم"}],
            [{"text": "🔙 بازگشت به مدرس"}],
        ],
        "resize_keyboard": True,
        "is_persistent": True,
    }

REGISTRATION_MENU = {
    "keyboard": [[{"text": "❌ لغو ثبت فرم"}], [{"text": "❌ انصراف"}]],
    "resize_keyboard": True,
    "one_time_keyboard": True,
}

EDIT_MENU = {
    "keyboard": [[{"text": "❌ لغو ویرایش"}], [{"text": "❌ انصراف"}]],
    "resize_keyboard": True,
    "one_time_keyboard": True,
}


def request_export_confirmation(action, user_id, chat_id):
    title = EXPORT_ACTIONS[action][0]
    set_state(
        f"pending_button_action:{user_id}",
        json.dumps({"action": action, "created_at": int(time.time())}),
    )
    markup = {
        "inline_keyboard": [[
            {"text": "✅ تأیید و ارسال", "callback_data": f"confirm:{action}"},
            {"text": "❌ لغو", "callback_data": "cancel"},
        ]]
    }
    send_message(
        f"⚠️ آیا از ساخت و ارسال «{title}» مطمئن هستید؟",
        chat_id=chat_id,
        reply_markup=markup,
    )


def execute_export_action(action):
    title, reports, include_processed = EXPORT_ACTIONS[action]
    if action == "all_full":
        reports = all_report_definitions("nursing")
    send_message(f"⏳ تأیید شد؛ در حال آماده‌سازی «{title}»…")
    result = run_report(
        include_processed=include_processed,
        selected_reports=list(reports),
    )
    log.info("Button action %s result: %s", action, result)


def request_custom_export_confirmation(mode, user_id, chat_id):
    selected = json.loads(get_state(f"selected_custom_form:{user_id}", "{}") or "{}")
    if not selected.get("survey_code") or not selected.get("report_name"):
        send_message("⚠️ ابتدا فرم موردنظر را از منوی اصلی انتخاب کنید.", chat_id, main_menu())
        return
    settings = get_form_settings(selected["survey_code"])
    self_issued = settings and settings["certificate_type"] == "self"
    title = (f"PowerPoint مدارک جدید فرم {selected['report_name']}" if self_issued
             else f"{'ثبت‌های جدید' if mode == 'new' else 'همه ثبت‌های'} فرم {selected['report_name']}")
    set_state(
        f"pending_button_action:{user_id}",
        json.dumps({
            "action": f"custom_{mode}",
            "survey_code": selected["survey_code"],
            "report_name": selected["report_name"],
            "certificate_type": settings["certificate_type"] if settings else "nursing",
            "created_at": int(time.time()),
        }, ensure_ascii=False),
    )
    markup = {"inline_keyboard": [[
        {"text": "✅ تأیید و ارسال", "callback_data": f"confirm:custom_{mode}"},
        {"text": "❌ لغو", "callback_data": "cancel"},
    ]]}
    send_message(f"⚠️ آیا از ساخت و ارسال «{title}» مطمئن هستید؟", chat_id, markup)


def execute_custom_export(pending):
    if pending.get("certificate_type") == "self":
        send_message(f"⏳ در حال ساخت PowerPoint دقیق مدارک فرم «{pending['report_name']}»…")
        result = run_certificate_export([pending["survey_code"]])
        log.info("Certificate export result: %s", result)
        return
    include_processed = pending["action"] == "custom_full"
    report = (pending["survey_code"], pending["report_name"])
    mode_title = "همه ثبت‌ها" if include_processed else "ثبت‌های جدید"
    send_message(f"⏳ در حال آماده‌سازی {mode_title}ی فرم «{pending['report_name']}»…")
    result = run_report(include_processed=include_processed, selected_reports=[report])
    log.info("Custom form export result: %s", result)


def form_new_counts(records):
    ids = resolve_surveys({code for _source, code, _name, _teacher in records})
    counts = {}
    for _source, code, _name, _teacher in records:
        settings = get_form_settings(code) or {"certificate_type": "nursing"}
        headers, rows, _total = fetch_results(ids[code])
        count = 0
        for raw_row in rows:
            key = response_key(row_to_mapping(headers, raw_row))
            done = (is_certificate_generated(code, key) if settings["certificate_type"] == "self"
                    else is_processed(code, key))
            if not done:
                count += 1
        counts[code] = count
    return counts


def send_status():
    send_message("⏳ در حال بررسی وضعیت همه فرم‌ها…")
    rows = run_status()
    if rows is None:
        send_message("⏳ یک گزارش دیگر در حال آماده‌سازی است. کمی بعد دوباره تلاش کنید.")
        return
    lines = ["📊 وضعیت فرم‌ها:"]
    for report_name, new_count, total in rows:
        lines.append(f"• {report_name}: 🆕 {new_count} جدید از 📚 {total} ثبت")
    send_message("\n".join(lines))


def normalize_digits(value):
    return str(value or "").translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789"))


def get_processed_times(survey_code, response_keys):
    if not response_keys:
        return {}
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT response_key, MAX(sent_at) FROM response_send_history
            WHERE survey_code=%s AND response_key=ANY(%s)
            GROUP BY response_key
            """,
            (survey_code, list(response_keys)),
        )
        return {row[0]: row[1] for row in cur.fetchall()}


def get_certificate_history(survey_code, response_keys):
    if not response_keys:
        return {}
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT i.response_key, b.short_id, b.created_at, b.nursing_received_at
            FROM report_batch_items i
            JOIN report_batches b ON b.id=i.batch_id
            WHERE i.survey_code=%s AND i.response_key=ANY(%s)
            ORDER BY b.created_at DESC
            """,
            (survey_code, list(response_keys)),
        )
        result = {}
        for key, short_id, created_at, received_at in cur.fetchall():
            result.setdefault(key, []).append({
                "short_id": short_id, "sent_at": created_at,
                "received_at": received_at,
            })
        return result


def get_self_certificate_history(survey_code, response_keys):
    if not response_keys:
        return {}
    with db() as conn, conn.cursor() as cur:
        cur.execute(
            """SELECT i.response_key, b.short_id, b.created_at
               FROM self_certificate_items i
               JOIN self_certificate_batches b ON b.id=i.batch_id
               WHERE i.survey_code=%s AND i.response_key=ANY(%s)
               ORDER BY b.created_at DESC""",
            (survey_code, list(response_keys)),
        )
        result = {}
        for key, short_id, created_at in cur.fetchall():
            result.setdefault(key, []).append({"short_id": short_id, "sent_at": created_at})
        return result


def registration_matches(person, query):
    normalized_query = normalize_digits(query).strip().casefold()
    if not normalized_query:
        return False
    national_id = normalize_digits(person.get("national_id", ""))
    if normalized_query.isdigit():
        return national_id == normalized_query
    names = f"{person.get('persian_name', '')} {person.get('english_name', '')}".casefold()
    return normalized_query in names


def gregorian_to_jalali(gy, gm, gd):
    g_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    j_days = [31, 31, 31, 31, 31, 31, 30, 30, 30, 30, 30, 29]
    gy -= 1600
    gm -= 1
    gd -= 1
    g_day_no = 365 * gy + (gy + 3) // 4 - (gy + 99) // 100 + (gy + 399) // 400
    for index in range(gm):
        g_day_no += g_days[index]
    if gm > 1 and ((gy + 1600) % 4 == 0 and ((gy + 1600) % 100 != 0 or (gy + 1600) % 400 == 0)):
        g_day_no += 1
    g_day_no += gd
    j_day_no = g_day_no - 79
    j_np = j_day_no // 12053
    j_day_no %= 12053
    jy = 979 + 33 * j_np + 4 * (j_day_no // 1461)
    j_day_no %= 1461
    if j_day_no >= 366:
        jy += (j_day_no - 1) // 365
        j_day_no = (j_day_no - 1) % 365
    jm = 0
    while jm < 11 and j_day_no >= j_days[jm]:
        j_day_no -= j_days[jm]
        jm += 1
    return jy, jm + 1, j_day_no + 1


def to_tehran_datetime(value):
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, (int, float)) or (isinstance(value, str) and normalize_digits(value).strip().isdigit()):
        number = float(normalize_digits(value).strip())
        if number > 10_000_000_000:
            number /= 1000
        if number < 1_000_000_000:
            return None
        parsed = datetime.fromtimestamp(number, tz=timezone.utc)
    else:
        text = normalize_digits(value).strip()
        if not text:
            return None
        iso_text = text.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(iso_text)
        except ValueError:
            parsed = None
            for pattern in (
                "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M",
                "%Y/%m/%d", "%Y-%m-%d",
            ):
                try:
                    parsed = datetime.strptime(text, pattern)
                    break
                except ValueError:
                    continue
            if parsed is None:
                return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=ZoneInfo(DISPLAY_TIMEZONE))
    return parsed.astimezone(ZoneInfo(DISPLAY_TIMEZONE))


def format_tehran_jalali(value):
    text = normalize_digits(value).strip() if not isinstance(value, datetime) else ""
    jalali_match = re.search(
        r"\b(1[34]\d{2})[-/](\d{1,2})[-/](\d{1,2})(?:[ T،]+(\d{1,2}):(\d{2})(?::\d{2})?)?",
        text,
    )
    if jalali_match:
        year, month, day = (int(jalali_match.group(i)) for i in range(1, 4))
        hour = int(jalali_match.group(4) or 0)
        minute = int(jalali_match.group(5) or 0)
        return f"{day:02d}-{month:02d}-{year:04d}، ساعت {hour:02d}:{minute:02d}"
    parsed = to_tehran_datetime(value)
    if parsed is None:
        return "نامشخص"
    jy, jm, jd = gregorian_to_jalali(parsed.year, parsed.month, parsed.day)
    return f"{jd:02d}-{jm:02d}-{jy:04d}، ساعت {parsed:%H:%M}"


def extract_submission_value(mapping):
    normalized = {clean_header(key): scalar_value(value) for key, value in mapping.items()}
    datetime_candidates = [
        "تاریخ و زمان تکمیل فرم", "تاریخ و ساعت تکمیل فرم",
        "submitted_at", "submission_date", "submit_date", "completed_at",
        "completion_date", "response_date", "created_at", "date_created",
    ]
    for candidate in datetime_candidates:
        value = normalized.get(clean_header(candidate))
        if value not in (None, ""):
            return str(value).strip()
    date_value = ""
    time_value = ""
    for candidate in (
        "تاریخ ثبت پاسخ", "تاریخ ارسال پاسخ", "تاریخ تکمیل",
        "تاریخ ثبت", "تاریخ پاسخ", "تاریخ ارسال", "تاریخ ایجاد", "date",
    ):
        value = normalized.get(clean_header(candidate))
        if value not in (None, ""):
            date_value = str(value).strip()
            break
    for candidate in (
        "زمان ثبت پاسخ", "زمان ارسال پاسخ", "زمان تکمیل",
        "زمان ثبت", "ساعت ثبت", "ساعت تکمیل", "ساعت ارسال", "time",
    ):
        value = normalized.get(clean_header(candidate))
        if value not in (None, ""):
            time_value = str(value).strip()
            break
    return " ".join(part for part in (date_value, time_value) if part)


def format_sent_at(value):
    return format_tehran_jalali(value)


def search_sent_registrations(query, limit=15):
    if not RUN_LOCK.acquire(blocking=False):
        return None, 0
    try:
        require_settings()
        init_db()
        records = get_form_records()
        ids = resolve_surveys({survey_code for _source, survey_code, _name, _teacher in records})
        matches = []
        for _source, survey_code, report_name, teacher_name in records:
            headers, rows, _total = fetch_results(ids[survey_code])
            candidates = []
            for raw_row in rows:
                mapping = row_to_mapping(headers, raw_row)
                person = extract_person(mapping)
                if registration_matches(person, query):
                    candidates.append((response_key(mapping), person, extract_submission_value(mapping)))
            processed = get_processed_times(survey_code, [key for key, _person, _submitted in candidates])
            certificates = get_certificate_history(
                survey_code, [key for key, _person, _submitted in candidates]
            )
            self_certificates = get_self_certificate_history(
                survey_code, [key for key, _person, _submitted in candidates]
            )
            for key, person, submitted_at in candidates:
                sent_at = processed.get(key)
                if not sent_at:
                    continue
                matches.append({
                    "person": person,
                    "form": report_name,
                    "teacher": teacher_name,
                    "sent_at": sent_at,
                    "submitted_at": submitted_at,
                    "certificate_reports": certificates.get(key, []),
                    "self_certificate_reports": self_certificates.get(key, []),
                })
        matches.sort(key=lambda item: item["sent_at"], reverse=True)
        return matches[:limit], len(matches)
    finally:
        RUN_LOCK.release()


def process_registration_search(text, user_id, chat_id):
    if text == "❌ لغو جست‌وجو":
        set_state(f"registration_search:{user_id}", "{}")
        send_message("❌ جست‌وجو لغو شد.", chat_id, main_menu())
        return
    query = text.strip()
    if len(query) < 2:
        send_message("⚠️ حداقل دو حرف از نام یا کد ملی کامل را وارد کنید.", chat_id, SEARCH_MENU)
        return
    send_message("🔍 در حال جست‌وجو در ثبت‌های ارسال‌شده…", chat_id=chat_id)
    matches, total_found = search_sent_registrations(query)
    if matches is None:
        send_message("⏳ عملیات دیگری در حال اجراست؛ کمی بعد دوباره امتحان کنید.", chat_id=chat_id)
        return
    set_state(f"registration_search:{user_id}", "{}")
    if not matches:
        send_message("❌ این نام یا کد ملی در ثبت‌های ارسال‌شده توسط ربات پیدا نشد.", chat_id, main_menu())
        return
    lines = [f"🔍 نتیجه جست‌وجوی «{query}»:"]
    for index, item in enumerate(matches, start=1):
        person = item["person"]
        lines.extend([
            "",
            f"{index}. 👤 {person['persian_name'] or person['english_name']}",
            f"🪪 کد ملی: {person['national_id'] or 'ثبت نشده'}",
            f"📋 فرم: {item['form']}",
            f"👨‍🏫 مدرس: {item['teacher']}",
            f"📝 تاریخ تکمیل فرم در پرس‌لاین: {format_tehran_jalali(item['submitted_at'])}",
            f"📤 آخرین ارسال توسط ربات: {format_sent_at(item['sent_at'])}",
        ])
        histories = item.get("certificate_reports", [])
        self_histories = item.get("self_certificate_reports", [])
        if not histories and not self_histories:
            lines.append("🎓 وضعیت مدرک: این ارسال مربوط به قبل از ثبت تاریخچه جدید است")
        for history in self_histories:
            lines.extend([
                f"🎓 گزارش {history['short_id']}: 🖨 PowerPoint ساخته و به چاپخانه ارسال شده",
                f"🕒 تاریخ وضعیت: {format_sent_at(history['sent_at'])}",
            ])
        for history in histories:
            if history["received_at"]:
                status = "✅ مدرک از نظام پرستاری تحویل گرفته شده"
                status_date = format_sent_at(history["received_at"])
            else:
                status = "📤 مدرک به نظام پرستاری ارسال شده؛ هنوز تحویل گرفته نشده"
                status_date = format_sent_at(history["sent_at"])
            lines.extend([
                f"🎓 گزارش {history['short_id']}: {status}",
                f"🕒 تاریخ وضعیت: {status_date}",
            ])
    if total_found > len(matches):
        lines.append(f"\nℹ️ {total_found - len(matches)} نتیجه دیگر نمایش داده نشد؛ جست‌وجو را دقیق‌تر کنید.")
    send_message("\n".join(lines), chat_id, main_menu())


def run_health_checks():
    checks = []
    started = time.monotonic()
    try:
        with db() as conn, conn.cursor() as cur:
            cur.execute("SELECT 1")
            cur.fetchone()
        checks.append(("پایگاه‌داده", True, time.monotonic() - started, "متصل"))
    except Exception as exc:
        checks.append(("پایگاه‌داده", False, time.monotonic() - started, type(exc).__name__))

    started = time.monotonic()
    try:
        porsline_get("/api/folders/")
        checks.append(("پرس‌لاین", True, time.monotonic() - started, "متصل"))
    except Exception as exc:
        checks.append(("پرس‌لاین", False, time.monotonic() - started, type(exc).__name__))

    started = time.monotonic()
    try:
        response = requests.get(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getMe",
            timeout=30,
        )
        response.raise_for_status()
        if not response.json().get("ok"):
            raise RuntimeError("Telegram rejected request")
        checks.append(("تلگرام", True, time.monotonic() - started, "متصل"))
    except Exception as exc:
        checks.append(("تلگرام", False, time.monotonic() - started, type(exc).__name__))
    return checks


def send_health_report(chat_id):
    send_message("🩺 در حال بررسی اتصال‌ها…", chat_id=chat_id)
    lines = ["🩺 وضعیت سلامت اتصال‌ها:"]
    for name, ok, elapsed, detail in run_health_checks():
        icon = "✅" if ok else "❌"
        lines.append(f"{icon} {name}: {detail} ({elapsed:.2f} ثانیه)")
    lines.append("\n🕒 زمان بررسی: " + format_sent_at(datetime.now(timezone.utc)))
    send_message("\n".join(lines), chat_id, main_menu())


def start_edit(kind, user_id, chat_id):
    if kind == "form":
        selected = json.loads(get_state(f"selected_custom_form:{user_id}", "{}") or "{}")
        if not selected.get("survey_code"):
            raise ValueError("⚠️ ابتدا یک فرم را انتخاب کنید.")
        payload = {"kind": "form", **selected}
        prompt = f"✏️ نام جدید فرم «{selected['report_name']}» را وارد کنید:"
    else:
        teacher_name = get_state(f"selected_teacher:{user_id}", "")
        if teacher_name not in get_teachers():
            raise ValueError("⚠️ ابتدا یک مدرس را انتخاب کنید.")
        payload = {"kind": "teacher", "teacher_name": teacher_name}
        prompt = f"✏️ نام جدید مدرس «{teacher_name}» را وارد کنید:"
    set_state(f"edit_entity:{user_id}", json.dumps(payload, ensure_ascii=False))
    send_message(prompt, chat_id, EDIT_MENU)


def process_edit(text, user_id, chat_id, editing):
    if text == "❌ لغو ویرایش":
        set_state(f"edit_entity:{user_id}", "{}")
        send_message("❌ ویرایش لغو شد.", chat_id, main_menu())
        return
    if editing["kind"] == "form":
        new_name = rename_form(editing["source"], editing["survey_code"], text)
        editing["report_name"] = new_name
        set_state(f"selected_custom_form:{user_id}", json.dumps(editing, ensure_ascii=False))
        teacher_name = editing["teacher_name"]
        send_message(f"✅ نام فرم به «{new_name}» تغییر کرد.", chat_id, teacher_menu(teacher_name))
    else:
        new_name = rename_teacher(editing["teacher_name"], text)
        set_state(f"selected_teacher:{user_id}", new_name)
        send_message(f"✅ نام مدرس به «{new_name}» تغییر کرد.", chat_id, teacher_menu(new_name))
    set_state(f"edit_entity:{user_id}", "{}")


def request_deactivation(kind, user_id, chat_id):
    if kind == "form":
        selected = json.loads(get_state(f"selected_custom_form:{user_id}", "{}") or "{}")
        if not selected.get("survey_code"):
            raise ValueError("⚠️ ابتدا یک فرم را انتخاب کنید.")
        payload = {"kind": "form", **selected}
        subject = f"فرم «{selected['report_name']}»"
    else:
        teacher_name = get_state(f"selected_teacher:{user_id}", "")
        if teacher_name not in get_teachers():
            raise ValueError("⚠️ ابتدا یک مدرس را انتخاب کنید.")
        payload = {"kind": "teacher", "teacher_name": teacher_name}
        subject = f"مدرس «{teacher_name}» و تمام فرم‌های او"
    payload.update({"stage": 1, "created_at": int(time.time())})
    set_state(f"pending_deactivation:{user_id}", json.dumps(payload, ensure_ascii=False))
    markup = {"inline_keyboard": [[
        {"text": "⚠️ تأیید مرحله اول", "callback_data": "deactivate:first"},
        {"text": "❌ لغو", "callback_data": "cancel"},
    ]]}
    send_message(
        f"⚠️ هشدار اول: {subject} از منو، وضعیت و گزارش‌ها غیرفعال می‌شود. ادامه می‌دهید؟",
        chat_id,
        markup,
    )


def process_deactivation_callback(data, user_id, chat_id):
    state_key = f"pending_deactivation:{user_id}"
    pending = json.loads(get_state(state_key, "{}") or "{}")
    if time.time() - float(pending.get("created_at", 0)) > 120:
        set_state(state_key, "{}")
        send_message("⌛ مهلت تأیید حذف تمام شده است؛ دوباره اقدام کنید.", chat_id=chat_id)
        return
    if data == "deactivate:first" and pending.get("stage") == 1:
        pending.update({"stage": 2, "created_at": int(time.time())})
        set_state(state_key, json.dumps(pending, ensure_ascii=False))
        subject = (f"فرم «{pending['report_name']}»" if pending["kind"] == "form"
                   else f"مدرس «{pending['teacher_name']}» و تمام فرم‌های او")
        markup = {"inline_keyboard": [[
            {"text": "🗑 تأیید نهایی غیرفعال‌سازی", "callback_data": "deactivate:second"},
            {"text": "❌ لغو", "callback_data": "cancel"},
        ]]}
        send_message(
            f"🚨 هشدار نهایی: آیا کاملاً مطمئن هستید که {subject} غیرفعال شود؟",
            chat_id,
            markup,
        )
        return
    if data != "deactivate:second" or pending.get("stage") != 2:
        send_message("⚠️ ترتیب تأیید معتبر نیست؛ دوباره اقدام کنید.", chat_id=chat_id)
        return
    if pending["kind"] == "form":
        changed = deactivate_form(pending["source"], pending["survey_code"])
        success = f"✅ فرم «{pending['report_name']}» غیرفعال شد. سوابق آن حفظ شده است."
    else:
        changed = deactivate_teacher(pending["teacher_name"])
        success = f"✅ مدرس «{pending['teacher_name']}» و تمام فرم‌هایش غیرفعال شدند. سوابق حفظ شده است."
    set_state(state_key, "{}")
    if not changed:
        raise ValueError("⚠️ مورد فعال برای غیرفعال‌سازی پیدا نشد.")
    send_message(success, chat_id, main_menu())


def process_restore_selection(text, chat_id):
    teacher_name = next(
        (name for name in get_inactive_teachers() if text == f"♻️ مدرس • {name}"),
        None,
    )
    if teacher_name:
        if not restore_teacher(teacher_name):
            raise ValueError("⚠️ مدرس غیرفعال موردنظر پیدا نشد.")
        send_message(
            f"✅ مدرس «{teacher_name}» همراه فرم‌هایش دوباره فعال شد.",
            chat_id,
            MANAGEMENT_MENU,
        )
        return True
    form = next(
        ((source, code, name, teacher) for source, code, name, teacher in get_inactive_form_records()
         if text == f"♻️ فرم • {name[:42]} • {str(code)[-8:]}"),
        None,
    )
    if form:
        source, survey_code, report_name, teacher_name = form
        if not restore_form(source, survey_code, teacher_name):
            raise ValueError("⚠️ فرم غیرفعال موردنظر پیدا نشد.")
        send_message(
            f"✅ فرم «{report_name}» و مدرس مربوط به آن دوباره فعال شدند.",
            chat_id,
            MANAGEMENT_MENU,
        )
        return True
    return False


def start_form_transfer(text, user_id, chat_id, form_records):
    selected = next(
        ((source, code, name, teacher) for source, code, name, teacher in form_records
         if text == f"🔄 {name[:46]} • {str(code)[-8:]}"),
        None,
    )
    if not selected:
        return False
    source, survey_code, report_name, teacher_name = selected
    targets = [name for name in get_teachers() if name != teacher_name]
    if not targets:
        send_message("⚠️ مدرس دیگری برای انتقال وجود ندارد.", chat_id, MANAGEMENT_MENU)
        return True
    set_state(
        f"form_transfer:{user_id}",
        json.dumps({
            "source": source,
            "survey_code": survey_code,
            "report_name": report_name,
            "teacher_name": teacher_name,
        }, ensure_ascii=False),
    )
    send_message(
        f"🔄 فرم «{report_name}» به کدام مدرس منتقل شود؟",
        chat_id,
        transfer_teacher_menu(teacher_name),
    )
    return True


def process_transfer_target(text, user_id, chat_id, transfer):
    if text == "❌ لغو انتقال":
        set_state(f"form_transfer:{user_id}", "{}")
        send_message("❌ انتقال فرم لغو شد.", chat_id, MANAGEMENT_MENU)
        return
    teacher_name = text.removeprefix("👨‍🏫 ").strip()
    if teacher_name not in get_teachers() or teacher_name == transfer["teacher_name"]:
        send_message("⚠️ یک مدرس مقصد معتبر انتخاب کنید.", chat_id, transfer_teacher_menu(transfer["teacher_name"]))
        return
    if not move_form(transfer["source"], transfer["survey_code"], teacher_name):
        raise ValueError("⚠️ فرم فعال موردنظر برای انتقال پیدا نشد.")
    set_state(f"form_transfer:{user_id}", "{}")
    send_message(
        f"✅ فرم «{transfer['report_name']}» از «{transfer['teacher_name']}» به «{teacher_name}» منتقل شد.",
        chat_id,
        MANAGEMENT_MENU,
    )


def process_form_registration(text, user_id, chat_id, registration):
    if text == "❌ لغو ثبت فرم":
        set_state(f"form_registration:{user_id}", "{}")
        send_message("❌ ثبت فرم لغو شد.", chat_id, main_menu())
        return
    if registration.get("step") == "name":
        report_name = text.strip()
        if not report_name or len(report_name) > 80:
            send_message("⚠️ نام فرم باید بین ۱ تا ۸۰ کاراکتر باشد.", chat_id, REGISTRATION_MENU)
            return
        set_state(
            f"form_registration:{user_id}",
            json.dumps({"step": "identifier", "report_name": report_name}, ensure_ascii=False),
        )
        send_message(
            "🔗 حالا شناسه فرم، آیدی عددی یا لینک کامل فرم در پرس‌لاین را بفرستید:",
            chat_id,
            REGISTRATION_MENU,
        )
        return
    if registration.get("step") == "identifier":
        survey_code = parse_survey_identifier(text)
        if survey_code in {str(code) for code, _name in all_report_definitions()}:
            raise ValueError("⚠️ این فرم قبلاً در ربات ثبت شده است.")
        send_message("🔍 در حال بررسی دسترسی به فرم پرس‌لاین…", chat_id=chat_id)
        validate_survey_access(survey_code)
        set_state(
            f"form_registration:{user_id}",
            json.dumps({
                "step": "teacher",
                "report_name": registration["report_name"],
                "survey_code": survey_code,
            }, ensure_ascii=False),
        )
        send_message("👨‍🏫 این فرم مربوط به کدام مدرس است؟", chat_id, teacher_selection_menu())
        return
    if registration.get("step") == "teacher":
        if text == "➕ مدرس جدید":
            registration["step"] = "new_teacher"
            set_state(
                f"form_registration:{user_id}",
                json.dumps(registration, ensure_ascii=False),
            )
            send_message("📝 نام مدرس جدید را وارد کنید:", chat_id, REGISTRATION_MENU)
            return
        teacher_name = text.removeprefix("👨‍🏫 ").strip()
        if teacher_name not in get_teachers():
            send_message("⚠️ لطفاً یکی از مدرس‌های فهرست را انتخاب کنید.", chat_id, teacher_selection_menu())
            return
        registration.update({"step": "certificate_type", "teacher_name": teacher_name})
        set_state(f"form_registration:{user_id}", json.dumps(registration, ensure_ascii=False))
        send_message("🎓 مدرک این فرم چگونه صادر می‌شود؟", chat_id, CERTIFICATE_TYPE_MENU)
        return
    if registration.get("step") == "new_teacher":
        if text.strip() in get_teachers():
            registration["step"] = "teacher"
            set_state(f"form_registration:{user_id}", json.dumps(registration, ensure_ascii=False))
            send_message(
                "⚠️ این مدرس قبلاً ثبت شده است؛ لطفاً او را از فهرست انتخاب کنید.",
                chat_id,
                teacher_selection_menu(),
            )
            return
        teacher_name = str(text or "").strip()
        if not teacher_name or len(teacher_name) > 60:
            send_message("⚠️ نام مدرس باید بین ۱ تا ۶۰ کاراکتر باشد.", chat_id, REGISTRATION_MENU)
            return
        registration.update({"step": "new_teacher_color", "teacher_name": teacher_name})
        set_state(f"form_registration:{user_id}", json.dumps(registration, ensure_ascii=False))
        send_message(f"🎨 رنگ مدرس «{teacher_name}» را انتخاب کنید:", chat_id, color_menu())
        return
    if registration.get("step") == "new_teacher_color":
        label = text.removeprefix("🎨 ").strip()
        color_key = next((key for key, name, _rgb in available_teacher_colors() if name == label), None)
        if not color_key:
            send_message("⚠️ یک رنگ در دسترس را انتخاب کنید.", chat_id, color_menu())
            return
        teacher_name = save_teacher(registration["teacher_name"], user_id, color_key)
        registration.update({"step": "certificate_type", "teacher_name": teacher_name})
        set_state(f"form_registration:{user_id}", json.dumps(registration, ensure_ascii=False))
        send_message("🎓 مدرک این فرم چگونه صادر می‌شود؟", chat_id, CERTIFICATE_TYPE_MENU)
        return
    if registration.get("step") == "certificate_type":
        if text == "🏥 برای نظام پرستاری":
            registration["certificate_type"] = "nursing"
            finish_form_registration(registration, registration["teacher_name"], user_id, chat_id)
            return
        if text != "🎓 خودم مدرک صادر می‌کنم":
            send_message("⚠️ یکی از دو نوع مدرک را انتخاب کنید.", chat_id, CERTIFICATE_TYPE_MENU)
            return
        registration.update({"step": "course_title", "certificate_type": "self"})
        set_state(f"form_registration:{user_id}", json.dumps(registration, ensure_ascii=False))
        send_message("📘 نام دوره‌ای که روی مدرک نوشته شود را وارد کنید:", chat_id, REGISTRATION_MENU)
        return
    prompts = {
        "course_title": ("course_duration", "⏱ مدت دوره را دقیقاً همان‌طور که باید روی مدرک نوشته شود وارد کنید:"),
        "course_duration": ("certificate_instructor", "👨‍🏫 نام مدرس روی مدرک را وارد کنید:"),
        "certificate_instructor": ("certificate_venue", "📍 محل برگزاری را وارد کنید:"),
        "certificate_venue": ("certificate_organization", "🏫 نام مجموعه آموزشی را وارد کنید:"),
    }
    if registration.get("step") in prompts:
        value = str(text or "").strip()
        if not value or len(value) > 150:
            send_message("⚠️ این مقدار باید بین ۱ تا ۱۵۰ کاراکتر باشد.", chat_id, REGISTRATION_MENU)
            return
        current = registration["step"]
        next_step, prompt = prompts[current]
        registration[current] = value
        registration["step"] = next_step
        set_state(f"form_registration:{user_id}", json.dumps(registration, ensure_ascii=False))
        send_message(prompt, chat_id, REGISTRATION_MENU)
        return
    if registration.get("step") == "certificate_organization":
        value = str(text or "").strip()
        if not value or len(value) > 150:
            send_message("⚠️ نام مجموعه آموزشی باید بین ۱ تا ۱۵۰ کاراکتر باشد.", chat_id, REGISTRATION_MENU)
            return
        registration["certificate_organization"] = value
        finish_form_registration(registration, registration["teacher_name"], user_id, chat_id)


def finish_form_registration(registration, teacher_name, user_id, chat_id):
        report_name = registration["report_name"]
        survey_code = registration["survey_code"]
        if registration.get("editing_settings"):
            update_form_certificate_settings(registration["source"], survey_code, registration)
        else:
            save_custom_form(
                custom_form_command(survey_code), survey_code, report_name, user_id, teacher_name, registration
            )
        set_state(f"form_registration:{user_id}", "{}")
        send_message(
            f"✅ مشخصات فرم «{report_name}» با موفقیت ذخیره شد.\n"
            + ("نوع مدرک: ساخت PowerPoint توسط ربات" if registration.get("certificate_type") == "self"
               else "نوع مدرک: نظام پرستاری"),
            chat_id,
            main_menu(),
        )


def cancel_all_operations(user_id, chat_id, notify_user=True):
    for prefix in (
        "form_registration", "edit_entity", "registration_search", "form_transfer",
        "teacher_setup", "multi_form", "batch_action", "pending_button_action",
        "pending_deactivation", "pending_batch",
    ):
        set_state(f"{prefix}:{user_id}", "{}")
    if notify_user:
        send_message("❌ عملیات جاری لغو شد.", chat_id, main_menu())


def process_teacher_setup(text, user_id, chat_id, setup):
    if setup.get("step") == "name":
        name = text.strip()
        if not name or len(name) > 60:
            send_message("⚠️ نام مدرس باید بین ۱ تا ۶۰ کاراکتر باشد.", chat_id=chat_id)
            return
        if name in get_teachers(active_only=False):
            raise ValueError("⚠️ این مدرس قبلاً ثبت شده است.")
        setup.update({"step": "color", "teacher_name": name})
        set_state(f"teacher_setup:{user_id}", json.dumps(setup, ensure_ascii=False))
        send_message(f"🎨 رنگ مدرس «{name}» را انتخاب کنید:", chat_id, color_menu())
        return
    if setup.get("step") == "select_teacher":
        teacher_name = text.removeprefix("🎨 مدرس • ").strip()
        if teacher_name not in get_teachers():
            send_message("⚠️ یک مدرس معتبر انتخاب کنید.", chat_id, teacher_color_selection_menu())
            return
        setup.update({"step": "edit_color", "teacher_name": teacher_name})
        set_state(f"teacher_setup:{user_id}", json.dumps(setup, ensure_ascii=False))
        send_message(f"🎨 رنگ جدید مدرس «{teacher_name}» را انتخاب کنید:", chat_id, color_menu(teacher_name))
        return
    if setup.get("step") in {"color", "edit_color"}:
        label = text.removeprefix("🎨 ").strip()
        available = available_teacher_colors(setup.get("teacher_name") if setup["step"] == "edit_color" else None)
        color_key = next((key for key, name, _rgb in available if name == label), None)
        if not color_key:
            send_message("⚠️ این رنگ در دسترس نیست؛ یکی از دکمه‌ها را انتخاب کنید.", chat_id, color_menu(setup.get("teacher_name")))
            return
        teacher_name = setup["teacher_name"]
        if setup["step"] == "color":
            save_teacher(teacher_name, user_id, color_key)
            message = f"✅ مدرس «{teacher_name}» با رنگ «{label}» اضافه شد."
        else:
            set_teacher_color(teacher_name, color_key)
            message = f"✅ رنگ مدرس «{teacher_name}» به «{label}» تغییر کرد."
        set_state(f"teacher_setup:{user_id}", "{}")
        send_message(message, chat_id, MANAGEMENT_MENU)


def start_batch_action(action, user_id, chat_id):
    stage = {"receive": "receive", "undo": "undo"}[action]
    batches, menu = batch_selection_menu(stage)
    if not batches:
        messages = {
            "receive": "گزارشی در انتظار تحویل از نظام پرستاری نیست.",
            "undo": "وضعیتی برای بازگردانی وجود ندارد.",
        }
        send_message("ℹ️ " + messages[action], chat_id, NURSING_MENU)
        return
    set_state(f"batch_action:{user_id}", json.dumps({"action": action}, ensure_ascii=False))
    instructions = {
        "receive": (
            "1️⃣ ثبت دریافت مدارک از نظام پرستاری\n\n"
            "فقط گزارشی را انتخاب کنید که مدارک تمام افراد آن به دست شما رسیده است. "
            "بعد از انتخاب، یک تأیید نهایی هم می‌گیرم."
        ),
        "undo": (
            "↩️ اصلاح انتخاب اشتباه\n\n"
            "گزارشی را که وضعیتش اشتباه ثبت شده انتخاب کنید؛ در مرحله بعد مشخص می‌کنید به کدام وضعیت برگردد."
        ),
    }
    send_message(instructions[action] + "\n\nیکی از گزارش‌های زیر را انتخاب کنید:", chat_id, menu)


def process_batch_selection(text, user_id, chat_id, action_state):
    if text == "🔙 بازگشت به مدارک":
        set_state(f"batch_action:{user_id}", "{}")
        send_message("🎓 مدیریت مدارک نظام پرستاری:", chat_id, NURSING_MENU)
        return
    action = action_state["action"]
    stage = {"receive": "receive", "undo": "undo"}[action]
    batch = next((item for item in recent_batches(stage) if text == batch_button_label(item)), None)
    if not batch:
        _batches, menu = batch_selection_menu(stage)
        send_message("⚠️ یکی از گزارش‌های فهرست را انتخاب کنید.", chat_id, menu)
        return
    set_state(f"batch_action:{user_id}", "{}")
    set_state(
        f"pending_batch:{user_id}",
        json.dumps({"action": action, "batch_id": batch["id"], "created_at": int(time.time())}),
    )
    if action == "undo":
        buttons = [[{"text": "↩️ بازگشت به «ارسال‌شده به نظام»", "callback_data": "batch:undo_receive"}]]
        buttons.append([{"text": "❌ انصراف", "callback_data": "cancel"}])
        send_message(f"وضعیت گزارش {batch['short_id']} به کدام مرحله برگردد؟", chat_id, {"inline_keyboard": buttons})
        return
    title = "تحویل‌گرفته‌شده از نظام پرستاری"
    markup = {"inline_keyboard": [[
        {"text": "✅ تأیید نهایی", "callback_data": f"batch:{action}"},
        {"text": "❌ انصراف", "callback_data": "cancel"},
    ]]}
    send_message(
        f"⚠️ تأیید نهایی\n\nگزارش: {batch['short_id']}\n"
        f"تعداد افراد: {batch['row_count']} نفر\n"
        f"فرم‌ها: {'، '.join(batch['report_names'])}\n\n"
        f"با تأیید شما، وضعیت تمام افراد این گزارش به «{title}» تغییر می‌کند. مطمئن هستید؟",
        chat_id, markup,
    )


def process_private_message(text, user_id, chat_id):
    try:
        if text == "❌ انصراف":
            cancel_all_operations(user_id, chat_id)
            return

        teacher_setup = json.loads(get_state(f"teacher_setup:{user_id}", "{}") or "{}")
        if teacher_setup.get("step"):
            process_teacher_setup(text, user_id, chat_id, teacher_setup)
            return

        batch_action = json.loads(get_state(f"batch_action:{user_id}", "{}") or "{}")
        if batch_action.get("action"):
            process_batch_selection(text, user_id, chat_id, batch_action)
            return

        registration = json.loads(get_state(f"form_registration:{user_id}", "{}") or "{}")
        if registration.get("step"):
            process_form_registration(text, user_id, chat_id, registration)
            return

        editing = json.loads(get_state(f"edit_entity:{user_id}", "{}") or "{}")
        if editing.get("kind"):
            process_edit(text, user_id, chat_id, editing)
            return

        searching = json.loads(get_state(f"registration_search:{user_id}", "{}") or "{}")
        if searching.get("active"):
            process_registration_search(text, user_id, chat_id)
            return

        transfer = json.loads(get_state(f"form_transfer:{user_id}", "{}") or "{}")
        if transfer.get("survey_code"):
            process_transfer_target(text, user_id, chat_id, transfer)
            return

        form_records = get_form_records()
        teachers = get_teachers()
        selected_teacher = next(
            (name for name in teachers if text == f"👨‍🏫 {name}"),
            None,
        )
        selected_custom = next(
            ((source, survey_code, report_name, teacher_name)
             for source, survey_code, report_name, teacher_name in form_records
             if text == custom_form_button_label(report_name, survey_code)),
            None,
        )
        if text == "⚙️ مدیریت ربات":
            send_message("⚙️ بخش مدیریت ربات:", chat_id, MANAGEMENT_MENU)
        elif text == "✖️ بستن منو":
            cancel_all_operations(user_id, chat_id, notify_user=False)
            send_message(
                "منوی ربات بسته شد. برای نمایش دوباره، /start را بفرستید.",
                chat_id,
                {"remove_keyboard": True},
            )
        elif text == "☑️ انتخاب چند فرم":
            send_message("⏳ در حال شمارش ثبت‌های جدید فرم‌ها…", chat_id=chat_id)
            records_for_count = get_form_records()
            counts = form_new_counts(records_for_count)
            state = {"selected": [], "stage": "select", "counts": counts}
            set_state(f"multi_form:{user_id}", json.dumps(state))
            send_message("☑️ حداقل دو فرم را انتخاب کنید؛ فرم با عدد صفر قابل انتخاب نیست:",
                         chat_id, multi_form_menu(counts=counts))
        elif text == "🎓 مدارک نظام پرستاری":
            send_message(nursing_help_text() + "\n\nاز کدام مرحله می‌خواهید استفاده کنید؟", chat_id, NURSING_MENU)
        elif text == "📘 راهنمای این بخش":
            send_message(nursing_help_text(), chat_id, NURSING_MENU)
        elif text == "📊 وضعیت گزارش‌های اخیر":
            send_message(nursing_status_text(), chat_id, NURSING_MENU)
        elif text == "1️⃣ مدارک از نظام رسید":
            start_batch_action("receive", user_id, chat_id)
        elif text == "↩️ اصلاح انتخاب اشتباه":
            start_batch_action("undo", user_id, chat_id)
        elif text == "➕ افزودن مدرس":
            set_state(f"teacher_setup:{user_id}", json.dumps({"step": "name"}))
            send_message("📝 نام مدرس جدید را وارد کنید:", chat_id, REGISTRATION_MENU)
        elif text == "🎨 رنگ مدرس‌ها":
            set_state(f"teacher_setup:{user_id}", json.dumps({"step": "select_teacher"}))
            send_message("🎨 مدرس موردنظر را انتخاب کنید:", chat_id, teacher_color_selection_menu())
        elif text == "🔙 بازگشت به مدیریت":
            send_message("⚙️ بخش مدیریت ربات:", chat_id, MANAGEMENT_MENU)
        elif text == "♻️ بازیابی موارد غیرفعال":
            inactive_count = len(get_inactive_teachers()) + len(get_inactive_form_records())
            if inactive_count:
                send_message("♻️ موردی را که می‌خواهید بازیابی شود انتخاب کنید:", chat_id, archive_menu())
            else:
                send_message("✅ مورد غیرفعالی وجود ندارد.", chat_id, MANAGEMENT_MENU)
        elif text == "🔄 انتقال فرم بین مدرس‌ها":
            if form_records:
                send_message("🔄 فرم موردنظر برای انتقال را انتخاب کنید:", chat_id, transfer_form_menu())
            else:
                send_message("⚠️ فرم فعالی برای انتقال وجود ندارد.", chat_id, MANAGEMENT_MENU)
        elif process_restore_selection(text, chat_id):
            return
        elif start_form_transfer(text, user_id, chat_id, form_records):
            return
        elif text == "🔍 جست‌وجوی ثبت‌نام‌کنندگان":
            set_state(f"registration_search:{user_id}", json.dumps({"active": True}))
            send_message(
                "🔍 نام فارسی، نام انگلیسی یا کد ملی فرد را وارد کنید:\n\n"
                "نتیجه شامل فرم، مدرس و تاریخ ارسال توسط ربات خواهد بود.",
                chat_id,
                SEARCH_MENU,
            )
        elif text == "🩺 بررسی سلامت":
            send_health_report(chat_id)
        elif selected_teacher:
            set_state(f"selected_teacher:{user_id}", selected_teacher)
            send_message(
                f"👨‍🏫 فرم‌ها و دوره‌های مدرس «{selected_teacher}»:",
                chat_id,
                teacher_menu(selected_teacher),
            )
        elif text == "🔙 بازگشت به منوی اصلی":
            send_message("🏠 منوی اصلی:", chat_id, main_menu())
        elif text == "🔙 بازگشت به مدرس":
            teacher_name = get_state(f"selected_teacher:{user_id}", "")
            if teacher_name in teachers:
                send_message(f"👨‍🏫 منوی مدرس «{teacher_name}»:", chat_id, teacher_menu(teacher_name))
            else:
                send_message("🏠 منوی اصلی:", chat_id, main_menu())
        elif text == "➕ ثبت فرم جدید":
            set_state(f"form_registration:{user_id}", json.dumps({"step": "name"}))
            send_message("📝 نام فرم جدید را وارد کنید:", chat_id, REGISTRATION_MENU)
        elif text == "📊 وضعیت فرم‌ها":
            send_status()
        elif selected_custom:
            source, survey_code, report_name, teacher_name = selected_custom
            set_state(
                f"selected_custom_form:{user_id}",
                json.dumps({
                    "source": source,
                    "survey_code": survey_code,
                    "report_name": report_name,
                    "teacher_name": teacher_name,
                }, ensure_ascii=False),
            )
            send_message(f"📋 خروجی موردنظر برای فرم «{report_name}» را انتخاب کنید:",
                         chat_id, custom_form_menu(survey_code))
        elif text == "🎓 ساخت PowerPoint مدارک جدید":
            request_custom_export_confirmation("new", user_id, chat_id)
        elif text == "🆕 فقط ثبت‌های جدید":
            request_custom_export_confirmation("new", user_id, chat_id)
        elif text == "📚 همه ثبت‌ها":
            request_custom_export_confirmation("full", user_id, chat_id)
        elif text == "✏️ ویرایش نام فرم":
            start_edit("form", user_id, chat_id)
        elif text == "⚙️ ویرایش مشخصات مدرک":
            selected = json.loads(get_state(f"selected_custom_form:{user_id}", "{}") or "{}")
            settings = get_form_settings(selected.get("survey_code"))
            if not settings:
                raise ValueError("⚠️ ابتدا یک فرم را انتخاب کنید.")
            settings.update({"step": "certificate_type", "editing_settings": True})
            set_state(f"form_registration:{user_id}", json.dumps(settings, ensure_ascii=False))
            send_message("🎓 نوع صدور مدرک این فرم را انتخاب کنید:", chat_id, CERTIFICATE_TYPE_MENU)
        elif text == "✏️ ویرایش نام مدرس":
            start_edit("teacher", user_id, chat_id)
        elif text == "🗑 غیرفعال‌کردن فرم":
            request_deactivation("form", user_id, chat_id)
        elif text == "🗑 غیرفعال‌کردن مدرس و تمام فرم‌ها":
            request_deactivation("teacher", user_id, chat_id)
        elif text in BUTTON_ACTIONS:
            request_export_confirmation(BUTTON_ACTIONS[text], user_id, chat_id)
        else:
            send_message("ℹ️ لطفاً یکی از گزینه‌های منو را انتخاب کنید.", chat_id, main_menu())
    except ValueError as exc:
        send_message(str(exc), chat_id=chat_id)
    except Exception as exc:
        log.exception("Private menu action failed")
        try:
            send_message(friendly_error_message(exc), chat_id=chat_id)
        except Exception:
            log.exception("Could not send private menu error")


def process_multi_callback(data, user_id, chat_id):
    state_key = f"multi_form:{user_id}"
    state = json.loads(get_state(state_key, "{}") or "{}")
    records = get_form_records()
    if not state:
        state = {"selected": [], "stage": "select"}
    selected = list(state.get("selected", []))
    counts = {str(key): int(value) for key, value in (state.get("counts") or {}).items()}
    if data.startswith("multi:zero:"):
        send_message("ℹ️ این فرم ثبت جدیدی ندارد و قابل انتخاب نیست.", chat_id,
                     multi_form_menu(selected, counts))
        return
    if data.startswith("multi:toggle:"):
        digest = data.rsplit(":", 1)[1]
        code = next((code for _source, code, _name, _teacher in records
                     if custom_form_command(code).endswith(digest)), None)
        if not code:
            raise ValueError("⚠️ فرم انتخاب‌شده دیگر فعال نیست.")
        if code in selected:
            selected.remove(code)
        else:
            selected.append(code)
        state.update({"selected": selected, "stage": "select"})
        set_state(state_key, json.dumps(state, ensure_ascii=False))
        send_message(f"☑️ {len(selected)} فرم انتخاب شده است.", chat_id, multi_form_menu(selected, counts))
        return
    if data == "multi:all":
        selected = [code for _source, code, _name, _teacher in records if counts.get(code, 0) > 0]
        state.update({"selected": selected, "stage": "select"})
        set_state(state_key, json.dumps(state, ensure_ascii=False))
        send_message("✅ همه فرم‌های دارای ثبت جدید انتخاب شدند.", chat_id, multi_form_menu(selected, counts))
        return
    if data == "multi:clear":
        state.update({"selected": [], "stage": "select"})
        set_state(state_key, json.dumps(state, ensure_ascii=False))
        send_message("🧹 انتخاب‌ها پاک شدند.", chat_id, multi_form_menu(counts=counts))
        return
    chosen = [(code, name, teacher) for _source, code, name, teacher in records if code in selected]
    if data == "multi:continue":
        if len(chosen) < 2:
            send_message("⚠️ برای خروجی چندفرمی حداقل دو فرم انتخاب کنید.", chat_id, multi_form_menu(selected, counts))
            return
        types = {get_form_settings(code)["certificate_type"] for code, _name, _teacher in chosen}
        if len(types) != 1:
            send_message("⚠️ فرم‌های نظام پرستاری و فرم‌های صدور شخصی را جداگانه انتخاب کنید.",
                         chat_id, multi_form_menu(selected, counts))
            return
        state["certificate_type"] = next(iter(types))
        colors = get_teacher_colors()
        missing = sorted({teacher for _code, _name, teacher in chosen if not colors.get(teacher)})
        if missing:
            set_state(state_key, "{}")
            send_message(
                "⚠️ ابتدا از مدیریت ربات برای این مدرس‌ها رنگ تعیین کنید: " + "، ".join(missing),
                chat_id, MANAGEMENT_MENU,
            )
            return
        state.update({"stage": 1, "created_at": int(time.time())})
        set_state(state_key, json.dumps(state, ensure_ascii=False))
        names = "\n".join(f"• {name}" for _code, name, _teacher in chosen)
        markup = {"inline_keyboard": [[
            {"text": "⚠️ تأیید مرحله اول", "callback_data": "multi:first"},
            {"text": "❌ انصراف", "callback_data": "cancel"},
        ]]}
        send_message("⚠️ گزارش جدید این فرم‌ها ساخته شود؟\n" + names, chat_id, markup)
        return
    if time.time() - float(state.get("created_at", 0)) > 180:
        set_state(state_key, "{}")
        send_message("⌛ مهلت تأیید تمام شد؛ دوباره فرم‌ها را انتخاب کنید.", chat_id=chat_id)
        return
    if data == "multi:first" and state.get("stage") == 1:
        state.update({"stage": 2, "created_at": int(time.time())})
        set_state(state_key, json.dumps(state, ensure_ascii=False))
        names = "، ".join(name for _code, name, _teacher in chosen)
        if state.get("certificate_type") == "self":
            markup = {"inline_keyboard": [
                [{"text": "📚 یک فایل مشترک", "callback_data": "multi:cert_combined"}],
                [{"text": "📂 فایل جدا برای هر فرم", "callback_data": "multi:cert_separate"}],
                [{"text": "❌ انصراف", "callback_data": "cancel"}],
            ]}
            send_message(
                f"⚠️ مرحله دوم — مدارک «{names}» به چه شکل ساخته شوند؟",
                chat_id, markup,
            )
            return
        markup = {"inline_keyboard": [[
            {"text": "✅ ساخت دو فایل", "callback_data": "multi:second"},
            {"text": "❌ انصراف", "callback_data": "cancel"},
        ]]}
        send_message(
            f"🚨 تأیید نهایی: ثبت‌های جدید «{names}» در یک گزارش سفید و یک گزارش رنگی ساخته شوند؟",
            chat_id, markup,
        )
        return
    if data in {"multi:cert_combined", "multi:cert_separate"} and state.get("stage") == 2:
        mode = "combined" if data.endswith("combined") else "separate"
        state.update({"stage": 3, "output_mode": mode, "created_at": int(time.time())})
        set_state(state_key, json.dumps(state, ensure_ascii=False))
        mode_title = "یک فایل مشترک" if mode == "combined" else "فایل جداگانه برای هر فرم"
        markup = {"inline_keyboard": [[
            {"text": "✅ تأیید نهایی و ساخت", "callback_data": "multi:cert_build"},
            {"text": "❌ انصراف", "callback_data": "cancel"},
        ]]}
        send_message(f"🚨 تأیید نهایی: PowerPoint مدارک به صورت «{mode_title}» ساخته و ارسال شود؟",
                     chat_id, markup)
        return
    if data == "multi:cert_build" and state.get("stage") == 3:
        set_state(state_key, "{}")
        send_message("⏳ در حال ساخت PowerPoint دقیق مدارک…", chat_id=chat_id)
        run_certificate_export([code for code, _name, _teacher in chosen], state["output_mode"], chat_id)
        return
    if data == "multi:second" and state.get("stage") == 2:
        set_state(state_key, "{}")
        send_message("⏳ در حال ساخت نسخه سفید و رنگی گزارش ترکیبی…", chat_id=chat_id)
        run_report(selected_reports=[(code, name) for code, name, _teacher in chosen], combine=True)
        return
    send_message("⚠️ ترتیب تأیید معتبر نیست؛ دوباره اقدام کنید.", chat_id=chat_id)


def process_batch_callback(data, user_id, chat_id):
    state_key = f"pending_batch:{user_id}"
    pending = json.loads(get_state(state_key, "{}") or "{}")
    if not pending or time.time() - float(pending.get("created_at", 0)) > 180:
        set_state(state_key, "{}")
        send_message("⌛ مهلت تأیید تمام شد؛ دوباره گزارش را انتخاب کنید.", chat_id, NURSING_MENU)
        return
    batch = get_batch(pending["batch_id"])
    if not batch:
        raise ValueError("⚠️ گزارش موردنظر پیدا نشد.")
    action = data.removeprefix("batch:")
    if pending.get("action") == "undo":
        if action != "undo_receive":
            raise ValueError("⚠️ عملیات بازگردانی معتبر نیست.")
    elif action != pending.get("action"):
        raise ValueError("⚠️ عملیات تأییدشده با درخواست جاری مطابقت ندارد.")
    if action == "receive" and batch["nursing_received_at"] is not None:
        raise ValueError("⚠️ این گزارش قبلاً از نظام پرستاری تحویل گرفته شده است.")
    update_batch_status(batch["id"], action)
    set_state(state_key, "{}")
    messages = {
        "receive": (
            "✅ ثبت شد: مدارک تمام افراد این گزارش از نظام پرستاری به دست شما رسیده است.\n\n"
            "این وضعیت از این پس در جست‌وجوی افراد نمایش داده می‌شود."
        ),
        "undo_receive": "✅ اصلاح شد: گزارش دوباره به وضعیت «ارسال‌شده به نظام؛ هنوز تحویل نگرفته‌اید» برگشت.",
    }
    send_message(messages[action], chat_id, NURSING_MENU)


def process_callback(callback_id, data, user_id, chat_id):
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
            json={"callback_query_id": callback_id},
            timeout=20,
        ).raise_for_status()
        state_key = f"pending_button_action:{user_id}"
        pending = json.loads(get_state(state_key, "{}") or "{}")
        if data == "cancel":
            set_state(state_key, "{}")
            set_state(f"pending_deactivation:{user_id}", "{}")
            set_state(f"multi_form:{user_id}", "{}")
            set_state(f"pending_batch:{user_id}", "{}")
            send_message("❌ عملیات لغو شد.", chat_id, main_menu())
            return
        if data.startswith("multi:"):
            process_multi_callback(data, user_id, chat_id)
            return
        if data.startswith("batch:"):
            process_batch_callback(data, user_id, chat_id)
            return
        if data.startswith("deactivate:"):
            process_deactivation_callback(data, user_id, chat_id)
            return
        if not data.startswith("confirm:"):
            return
        action = data.split(":", 1)[1]
        valid_action = action in EXPORT_ACTIONS or action in {"custom_new", "custom_full"}
        if (not valid_action or pending.get("action") != action
                or time.time() - float(pending.get("created_at", 0)) > 120):
            send_message("⌛ مهلت تأیید تمام شده است؛ دوباره گزینه موردنظر را انتخاب کنید.", chat_id=chat_id)
            return
        set_state(state_key, "{}")
        if action in {"custom_new", "custom_full"}:
            execute_custom_export(pending)
        else:
            execute_export_action(action)
    except Exception as exc:
        log.exception("Callback action failed")
        try:
            send_message(friendly_error_message(exc), chat_id=chat_id)
        except Exception:
            log.exception("Could not send callback error")


def telegram_get_updates(offset):
    response = requests.get(
        f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates",
        params={"offset": offset, "timeout": 25, "allowed_updates": json.dumps(["message"])},
        timeout=35,
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("ok"):
        raise RuntimeError(f"Telegram getUpdates failed: {payload}")
    return payload.get("result", [])


def telegram_poll_loop():
    while True:
        try:
            require_settings()
            init_db()
            offset = int(get_state("telegram_update_offset", "0"))
            for update in telegram_get_updates(offset):
                update_id = int(update["update_id"])
                offset = update_id + 1
                set_state("telegram_update_offset", offset)
                message = update.get("message") or {}
                chat_id = str((message.get("chat") or {}).get("id", ""))
                text = str(message.get("text") or "").strip()
                command = text.split()[0].split("@")[0].lower() if text else ""
                if chat_id != str(TELEGRAM_CHAT_ID) or not command.startswith("/"):
                    continue
                user_id = str((message.get("from") or {}).get("id", "unknown"))
                process_command_safely(command, user_id, text)
        except Exception:
            log.exception("Telegram polling failed")
            time.sleep(10)


def handle_report_command(command, user_id="unknown", full_text=""):
    if command == "/help":
        send_message(help_text())
        return
    if command == "/forms":
        init_db()
        lines = ["فرم‌های فعال:"]
        for form_command, (_code, report_name) in SINGLE_REPORT_COMMANDS.items():
            lines.append(f"• {form_command} — {report_name}")
        for form_command, _code, report_name in get_custom_forms():
            lines.append(f"• {form_command} — {report_name}")
        send_message("\n".join(lines))
        return
    if command == "/add_form":
        if not is_group_admin(user_id):
            send_message("فقط مدیران گروه اجازه افزودن فرم جدید را دارند.")
            return
        init_db()
        new_command, survey_code, report_name = parse_add_form(full_text)
        if survey_code in {code for code, _name in SINGLE_REPORT_COMMANDS.values()}:
            send_message("این فرم از قبل در ربات ثبت شده است.")
            return
        resolve_surveys({survey_code})
        save_custom_form(new_command, survey_code, report_name, user_id)
        send_message(
            f"فرم «{report_name}» با موفقیت اضافه شد.\n"
            f"دستور دریافت پاسخ‌های جدید: {new_command}"
        )
        return
    if command == "/remove_form":
        if not is_group_admin(user_id):
            send_message("فقط مدیران گروه اجازه غیرفعال‌کردن فرم را دارند.")
            return
        parts = full_text.split()
        if len(parts) != 2 or not re.fullmatch(r"/[a-z][a-z0-9_]{1,30}", parts[1].lower()):
            send_message("فرمت درست:\n/remove_form /command")
            return
        target = parts[1].lower()
        if target in SINGLE_REPORT_COMMANDS:
            send_message("چهار فرم اصلی از داخل تلگرام قابل حذف نیستند.")
            return
        init_db()
        if deactivate_custom_form(target):
            send_message(f"فرم مربوط به دستور {target} غیرفعال شد.")
        else:
            send_message("فرم فعالی با این دستور پیدا نشد.")
        return
    if command == "/status":
        send_message("در حال بررسی وضعیت همه فرم‌ها…")
        rows = run_status()
        if rows is None:
            send_message("یک گزارش دیگر در حال آماده‌سازی است. کمی بعد دوباره تلاش کنید.")
            return
        lines = ["وضعیت فرم‌ها:"]
        for report_name, new_count, total in rows:
            lines.append(f"• {report_name}: {new_count} پاسخ جدید از مجموع {total}")
        send_message("\n".join(lines))
        return
    init_db()
    selected_report = get_report_for_command(command)
    if selected_report:
        report_name = selected_report[1]
        send_message(f"در حال آماده‌سازی گزارش {report_name}…")
        result = run_single_report(command)
        log.info("Single report command %s result: %s", command, result)
        return
    if command == "/report_all":
        init_db()
        set_state(f"report_all_confirmation:{user_id}", int(time.time()))
        send_message(
            "این دستور تمام پاسخ‌های همه فرم‌های فعال را دوباره ارسال می‌کند. "
            "برای تأیید، حداکثر تا دو دقیقه دستور /confirm_report_all را بفرستید. "
            "برای لغو /cancel_report را بفرستید."
        )
        return
    elif command == "/confirm_report_all":
        init_db()
        state_key = f"report_all_confirmation:{user_id}"
        requested_at = float(get_state(state_key, "0") or 0)
        if time.time() - requested_at > 120:
            send_message("درخواست تأیید وجود ندارد یا مهلت دو دقیقه‌ای آن تمام شده است. دوباره /report_all را بفرستید.")
            return
        set_state(state_key, "0")
        send_message("تأیید شد؛ در حال آماده‌سازی گزارش کامل همه فرم‌ها…")
        result = run_report(include_processed=True, all_forms=True)
    elif command == "/cancel_report":
        init_db()
        set_state(f"report_all_confirmation:{user_id}", "0")
        send_message("ارسال گزارش کامل لغو شد.")
        return
    elif command == "/report_all_new":
        send_message("در حال بررسی پاسخ‌های جدید همه فرم‌ها…")
        result = run_report(include_processed=False, all_forms=True)
    elif command == "/report":
        send_message("در حال آماده‌سازی دو گزارش جدید…")
        result = run_report()
    else:
        return
    log.info("Command report result: %s", result)


def process_command_safely(command, user_id="unknown", full_text=""):
    try:
        handle_report_command(command, user_id, full_text)
    except ValueError as exc:
        send_message(str(exc))
    except Exception as exc:
        log.exception("Telegram command %s failed", command)
        try:
            send_message(friendly_error_message(exc))
        except Exception:
            log.exception("Could not send friendly error message")


def register_telegram_webhook():
    if not WEBHOOK_BASE_URL or not WEBHOOK_SECRET:
        return
    response = requests.post(
        f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/setWebhook",
        json={
            "url": f"{WEBHOOK_BASE_URL}/telegram-webhook",
            "secret_token": WEBHOOK_SECRET,
            "allowed_updates": ["message", "callback_query"],
            "drop_pending_updates": True,
        },
        timeout=45,
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("ok"):
        raise RuntimeError(f"Telegram rejected webhook: {payload}")
    log.info("Telegram webhook registered")


@app.get("/")
@app.get("/health")
def health():
    return jsonify({"ok": True, "service": "porsline-telegram-reporter"})


@app.post("/telegram-webhook")
def telegram_webhook():
    if not WEBHOOK_SECRET or request.headers.get("X-Telegram-Bot-Api-Secret-Token") != WEBHOOK_SECRET:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    update = request.get_json(silent=True) or {}
    callback = update.get("callback_query")
    if callback:
        callback_message = callback.get("message") or {}
        callback_chat = callback_message.get("chat") or {}
        user_id = str((callback.get("from") or {}).get("id", ""))
        chat_id = str(callback_chat.get("id", ""))
        if callback_chat.get("type") != "private":
            return jsonify({"ok": True})
        init_db()
        if user_id != str(get_state("bot_owner_user_id", "")):
            return jsonify({"ok": True})
        threading.Thread(
            target=process_callback,
            args=(str(callback.get("id", "")), str(callback.get("data", "")), user_id, chat_id),
            daemon=True,
        ).start()
        return jsonify({"ok": True})

    message = update.get("message") or {}
    chat = message.get("chat") or {}
    if chat.get("type") != "private":
        return jsonify({"ok": True})
    chat_id = str(chat.get("id", ""))
    user_id = str((message.get("from") or {}).get("id", ""))
    text = str(message.get("text") or "").strip()
    init_db()
    registered_owner = str(get_state("bot_owner_user_id", ""))

    if not registered_owner:
        parts = text.split(maxsplit=1)
        supplied_code = parts[1].strip() if len(parts) == 2 and parts[0].lower() == "/start" else ""
        if not BOT_ACCESS_CODE:
            send_message("رمز ورود ربات هنوز در Render تنظیم نشده است.", chat_id=chat_id)
        elif supplied_code and hmac.compare_digest(supplied_code, BOT_ACCESS_CODE):
            set_state("bot_owner_user_id", user_id)
            set_state("bot_owner_chat_id", chat_id)
            send_message("✅ دسترسی اختصاصی شما فعال شد. از منوی زیر استفاده کنید:", chat_id, main_menu())
        else:
            send_message("برای فعال‌سازی، دستور /start را همراه رمز اختصاصی وارد کنید.", chat_id=chat_id)
        return jsonify({"ok": True})

    if user_id != registered_owner:
        send_message("شما اجازه استفاده از این ربات را ندارید.", chat_id=chat_id)
        return jsonify({"ok": True})

    set_state("bot_owner_chat_id", chat_id)
    if text and text.split(maxsplit=1)[0].lower() == "/start":
        send_message("🏠 منوی اصلی:", chat_id, main_menu())
    elif text:
        threading.Thread(target=process_private_message, args=(text, user_id, chat_id), daemon=True).start()
    return jsonify({"ok": True})


@app.post("/run-now")
def run_now():
    supplied = request.headers.get("X-App-Secret") or request.args.get("secret")
    if not APP_SECRET or supplied != APP_SECRET:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    try:
        return jsonify({"ok": True, **run_report()})
    except Exception as exc:
        log.exception("Manual report failed")
        return jsonify({"ok": False, "error": str(exc)}), 500


if TELEGRAM_POLLING_ENABLED and os.getenv("DISABLE_TELEGRAM_POLLING", "false").lower() != "true":
    threading.Thread(target=telegram_poll_loop, daemon=True).start()

if WEBHOOK_BASE_URL and WEBHOOK_SECRET:
    threading.Thread(target=register_telegram_webhook, daemon=True).start()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8080")))
